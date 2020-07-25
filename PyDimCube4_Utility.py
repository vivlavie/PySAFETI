#PyDimCube4_Utility
#Effect of process fire onto Utility modules
#For a cube, read from each cube the duration that jet fire impinges on it
#PyDimCube3
#To consider failure cases for fire or gas detector
#Failure of fire detector for immediate ignition
#Failure of gas detectin for delayed ignition

#Pool fire is neglected

from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill
import sys
import matplotlib.pylab as pltparam


# sys.stdout = open('DimCube2.txt','w')

# FireWallBtn3and4 = False
# FirewallBtn3and4X = 140.7 #meter from AP

FireWallBtnUandP = True
FirewallBtnUandPX = 98.7 #meter from AP

WantPlot = True
cube_result2file = True

DimCubeSuccessList = {}
DimCubeFailList = {}


def print_cum_cube(cube,AA):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube))
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F))
        if F>1.0E-3:
            break
def print_cum_cube_file(cube,AA,a_file):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube),file = a_file)
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F),file = a_file)
        if F>1.0E-3:
            break

import dill
# Area = "ProcessArea"
element_dump_filename = 'Bv06_dump'
# icubeloc='SCE_CUBE_XYZ2_Process'

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)
lEvent_P = lEvent

# Area = "HullDeck"
# element_dump_filename = 'Bv06_hull_dump'
# icubeloc='SCE_CUBE_XYZ2_HullDeck'

# element_dump_filename = 'Bv06_offloading_dump'
# icubeloc='SCE_CUBE_XYZ2_Offloading'

element_dump_filename = 'Bv06_utility_dump'
icubeloc='SCE_CUBE_XYZ2_Utility'

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)
lEvent_U = lEvent

lEvent = lEvent_P + lEvent_U

#Read 'Presure vessel' from Input
iExlFilename='Bv06_i'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
Z = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv    
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    r += 1
numPV = r-63
def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    if m>5:
        jf = -13.2+54.3*math.log10(m)
    elif m>0.1:
        jf= 3.736*m + 6.
    else:
        jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
iIS=load_workbook(filename='IS_v12_shk.xlsx')
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 82:
    pv = shIS.cell(r,5).value
    IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    Modules[pv] = shIS.cell(r,7).value
    Deck[pv] =  shIS.cell(r,8).value
    if shIS.cell(r,24).value != None:
       nedvs = shIS.cell(r,24).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1



# iExlFilename='025-02-01-G_i'
# iExl=load_workbook(filename=iExlFilename+'.xlsx')
# cExlFilename='025-02-01-G_c'

#Failure of fire detector for immediate ignition
P_FD_Fail = 0.05 #default value? #should be adjusted for hole size or leak rate? To be suggested to use the release rate!
# P_FD_Fail['SM'] = 0.1
# P_FD_Fail['ME'] = 0.05
# P_FD_Fail['MA'] = 0.005
# P_FD_Fail['LA'] = 0.005

#Failure of gas detectin for delayed ignition
P_GD_Fail = 0.05


iExl=load_workbook(filename=icubeloc+'.xlsx')
shCube = iExl['xyz']
ncube = shCube.cell(1,1).value
Xcube = {}
Ycube = {}
Zcube = {}
Cubes = []
CubeDeck = {}
CubeImpingeDuration = {}
for i in range(3,ncube+3):
    id = shCube.cell(i,1).value
    Cubes.append(id)
    Xcube[id] = shCube.cell(i,2).value
    Ycube[id] = shCube.cell(i,3).value
    Zcube[id] = shCube.cell(i,4).value
    CubeDeck[id] = shCube.cell(i,5).value

print("{:32s} {:5s} {:4s} {:10s} Duration[sec]  Duration[min] m_dot[kg/s] or Pool diameter[m]".\
                format('Scenario','Module','Deck','Cube'))
ImpingeDurationArray = []

for i in range(0,ncube):    
# for i in range(0,37):
# for i in [ncube-1]:
    # li = 0
    # sep = np.zeros((len(lEvent),2))
    ImpingeDurationArray = []
    id = Cubes[i]
    xx = Xcube[id]
    yy = Ycube[id]
    zz = Zcube[id]
    if cube_result2file == True:
        f_cube_result = open(id+"_P.txt","w")
    # if (zz >= 35.0) and (zz < 44):
    #     CubeDeck = 'A'
    # elif (zz >= 44.0) and (zz < 53):
    #     CubeDeck = 'B'
    # elif (zz >= 53):
    #     CubeDeck = 'C'
    # elif (zz < 35 ):
    #     CubeDeck = 'Hull'
    
    for ei in range(0,len(lEvent)):
        e = lEvent[ei]
        # print(id,e.Module, e.Key)
        if (CubeDeck[id] == e.Deck) or (id[0:3]) == e.Module:
        # if True:
            #Exposure to jet fire
            #read frequency
            # sep[li,1] = e.JetFire.Frequency
            #check if the receptor is in the flame
            dx = xx - e.X
            dy = yy - e.Y
            dz = zz - (e.ReleaseHeight+35)
            # rr = math.sqrt(dx*dx+dy*dy+dz*dz)
            rr = math.sqrt(dx*dx+dy*dy)
            #Array of jet length
            jl_e = e.jfscale*2.8893*np.power(55.5*e.TVD[:,2],0.3728) #the 3rd column of the matrix, Lowesmith formulat
            t_e = interp1d(jl_e,e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'            

            if e.Discharge.ReleaseRate <= 1:
                P_FD_Fail = 0.1
                P_GD_Fail = 0.1
            elif e.Discharge.ReleaseRate <= 10:
                P_FD_Fail = 0.05
                P_GD_Fail = 0.05
            elif e.Discharge.ReleaseRate > 10:
                P_FD_Fail = 0.005
                P_GD_Fail = 0.005
            else:
                print('something wrong in P_FD_Fail')

            # The following condition is to consider effects of fire wall
            # If the protection by firewall is to be considered, fire at a PORT process moduel will not affect E&I building but may do GTG.
            # If the protection by firewall is to be considered, fire at a STBD process moduel may affect E&I building but will not do GTG.            # 
            
            if (FireWallBtnUandP == True) and (((yy > 0) and (e.Y < 0)) or ((yy < 0) and (e.Y > 0))):
                if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                    #When ESD fails, whether it is dueto Fire detection failure should be dinstinguished
                    #ESD fail even upon Fire detection successful
                    #Probability of failure of ESD and BDV is already reflected in e.JetFire.Frequency???
                    ff = e.JetFire.Frequency*(1-P_FD_Fail)
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FO_Jet"])                    

                    #ESD is not activated due to fire detection failure, any release irrespective of success of ESDV and BDV will result in EXBX release case      
                    ff = e.JetFire.Frequency/e.PESD/e.PBDV*P_FD_Fail
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FX_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FX_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FX_Jet"])                    
                else:  #Fire detection succedded
                    #Probability of success of ESD is already reflected in e.JetFire.Frequency???
                    ff = e.JetFire.Frequency*(1-P_FD_Fail)
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FO_Jet"])                    
      
    #To pin-point a scenario that give the dimensioning scenario
    IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0]) #with the longest duration at the bottom
    
    cf = 0
    jp = IDAsorted[-1]
    DimFreq = 1.0E-4
    di = 0
    cfp = jp[1]#frequency for the longest duration jet fire
    InterpolationSuccess = False
    if cfp > DimFreq:
        di = jp[0]
        scn = jp[2]
    else:
        for j in IDAsorted[-2:0:-1]:
            cf = cfp + j[1]
            if cf >= DimFreq and cfp < DimFreq:
                dp = jp[0]
                dn = j[0]
                # di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
                # print('Dimensioning jet duration {:8.1f}'.format(di))
                scnp = jp[2]
                scn = j[2]
                InterpolationSuccess = True
                #Choose the scenario close to the threshold
                if abs(cf-DimFreq) > abs(cfp-DimFreq):
                    scn = scnp
                    di = dp
                else:
                    di = dn
                break
            cfp = cf
            # print(cfp,cf)
            jp = j
    
    # print(scn)
    if InterpolationSuccess:
        CubeImpingeDuration[Cubes[i]] = di
                
        pv,hole,weather = scn.split("\\")            
        weather_fire = weather            
        x1 = X[pv]
        y1 = Y[pv]
        z1 = Z[pv]
        for e in lEvent:
            if e.Key in scn:
                break
        rrifound = False
        rri = 0
        rr5min = 0.
        for i in range(0,len(e.TVD)-1):
            if (e.TVD[i,0] < di) and (e.TVD[i+1,0] >= di):
                rri = e.TVD[i,2]
                rr5min = e.Discharge.RRs[1]
                rrifound = True
                break
        if rrifound == False:
            print(id,"Something wrong, rri not found", di, e.TVD[-1,0])

        # Max width of jet fire and its positing along 'x' (61%)
        # z1 = z1+35
        # dx=xx-x1
        # dy=yy-y1
        # dz=zz-z1
        # jli = jffit(rri)
        # xm = x1 + 0.61*dx
        # ym = y1 + 0.61*dy
        # zm = z1 + 0.61*dz
        # ll = math.sqrt(dx*dx+dy*dy+dz*dz)
        # print("IS: {:35s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:8.1f}{:8.1f}".\
        #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,di,rri))
        print("{:35s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} {:8.1f}".\
            format(scn, Modules[pv],Deck[pv],id,di,di/60,rri,rr5min))
        DimCubeSuccessList[id] = "'tvd-"+pv+"_"+hole
        if cube_result2file == True:                
            print_cum_cube_file(id,IDAsorted,f_cube_result) 

        # print("IS: {:35s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:4s}{:2s} -> {:8s}{:8.1f}{:8.1f}{:8.1f}".\
        #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,Modules[pv],Deck[pv],id,di,rri,2.8893*np.power(55.5*rri,0.3728)))
        # print("IS: {:35s} {:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Release rate: {:8.1f}".\
            # format(scn,x1,y1,z1,id,xx,yy,zz,rri))
        # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*x1,1000*y1,1000*z1,1000*xm,1000*ym,1000*zm,0,0.12*jli*1000))
        # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*xm,1000*ym,1000*zm,1000*xx,1000*yy,1000*zz,0.12*jli*1000,0))
        # print_cum_cube(id,IDAsorted) 
    elif not InterpolationSuccess:
        print(" No dimensioning scenario",id)
        scn = 'No dimensioning scenario for ' + id
        # print_cum_cube(id+" No dimensioning scenario ",IDAsorted)        
        if cube_result2file == True:
            print_cum_cube_file(id+" No dimensioning scenario ",IDAsorted,f_cube_result)        


    # if (InterpolationSuccess == True) and (WantPlot == True):
    if (WantPlot == True):
        ll = len(IDAsorted)
        ec = np.zeros([ll,2])
        i=0
        ec[i,1] = IDAsorted[-1][0] #the longest duration
        ec[i,0] = IDAsorted[-1][1] #Frequency for the longest duration
        for i in range(1,len(IDAsorted)):
            ec[i,1] = IDAsorted[ll-i-1][0]
            ec[i,0] = ec[i-1,0] + IDAsorted[ll-i-1][1]

        # plt.figure(figsize=(5.91, 3.15))
        CF = ec[1:,0]
        JFL = ec[1:,1]
        masscolor = 'tab:blue'
        fig,ax1 = plt.subplots()
        ax1.set_xlabel('Jet Impingement Duration [sec]')
        ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
        ax1.semilogy(JFL,CF,color=masscolor)
        ax1.set_ylim(top=5E-4,bottom=1E-6)
        ax1.set_xlim(left=0, right=3600)
        ax1.tick_params(axis='y',labelcolor=masscolor)
        ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
        ax1.annotate(scn,xy=(di,1.0E-4),xytext=(di,2E-4),horizontalalignment='left',verticalalignment='top',arrowprops = dict(facecolor='black',headwidth=4,width=2,headlength=4))
        # ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
        # ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
        # ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
        ax1.grid(True,which="major")

        if InterpolationSuccess:
            plt.title("Cube {:10s} - {:5s} fire from {:30} for {:8.1f} [sec]".format(id,weather_fire[6:],pv+"_"+hole+"_"+weather_fire[:5],di))            
        else:
            plt.title("Cube {:10s} - No dimensioning fire".format(id))            
        plt.tight_layout()
        plt.show()

        fig.savefig("{}_P.png".format(id))
        plt.close()

    if cube_result2file == True:
        f_cube_result.close()
# print_cum_cube(IDAsorted)

# for c in Cubes:
#     print("{:20s} {:8.1f} [min]".format(c,CubeImpingeDuration[c]/60))
# sys.stdout.close()

# for s in ['S03','S04','S05']:
#     hazcount = 0
#     print( "{:40s}{:10s} <> {:9}".format(s,"Jet Length","Distance"))
#     for e in lEvent:
#         if (s in e.Module):
#             if ((e.X-xx) < e.JetFire.Length):
#                 print( "{:40s}{:10.2f} > {:10.2f}".format(e.Key,e.JetFire.Length,e.X-xx))
#                 hazcount += 1
#             # else:
#                 # print( "{:40s}{:10.2f} < {:10.2f}".format(e.Key,e.JetFire.Length,e.X-xx))       
#     print(hazcount)
