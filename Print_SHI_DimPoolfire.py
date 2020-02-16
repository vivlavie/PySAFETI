from openpyxl import load_workbook
import math
import numpy as np
#import dill

iExlFilename='h11c4_dim_pool_fire'
Modules = ['CDA','KOD','MeOHS','HFP','HFS','HAP','HAS']
Area = "Hull"


iExlFilename='v04c_dim_pool_fire'
Modules = ['P02','P03','P04','P05','S02','S03','S04','S05']
Area = "Topside"
# element_dump_filename = 'v09_dump_jet'

iExl=load_workbook(filename=iExlFilename+'.xlsx')
shtPool = iExl['Sheet1']

# with open(element_dump_filename,'rb') as elements_dump:
#     lEvent = dill.load(elements_dump)  
def DtoH (D05):
    if D05 > 19.7:
            H05 = 0.4664*D05 + 18.345
    elif D05 < 1.0:
        H05 = 0.
    else: #  1.0 < D05 < 19.7:
        H05 = 1.2592*D05 + 2.7235
    return H05

#EPSEP  & LPSEP should have been read
#for r in range(6,numscn):
numrows = shtPool.max_row
Dxx = []
FreqTotal = 0.
for r in range(2,numrows+1):
    module = shtPool.cell(r,1).value
    path = shtPool.cell(r,2).value
    if Area == "Hull":
        area, pv, eq_hole_weather = path.split("\\")
    elif Area == "Topside":
        pv, eq_hole_weather = path.split("\\")
        pv = "0"+pv
        area = Area
    weather = eq_hole_weather[-5:]
    hole = eq_hole_weather[-8:-6]
    eq = eq_hole_weather[:-9]
    
    rr = shtPool.cell(r,3).value
    t2 = shtPool.cell(r,4).value
    mass_split = shtPool.cell(r,5).value
    burn_rate = shtPool.cell(r,6).value
    D0 = shtPool.cell(r,7).value
    H0 = shtPool.cell(r,8).value
    Te = shtPool.cell(r,9).value
    D00 = shtPool.cell(r,10).value
    D05 = shtPool.cell(r,11).value
    D10 = shtPool.cell(r,12).value
    D30 = shtPool.cell(r,13).value
    D60 = shtPool.cell(r,14).value
    H00 = H0
    H05 = DtoH(D05)
    H10 = DtoH(D10)
    H30 = DtoH(D30)
    H60 = DtoH(D60)

    Freq = shtPool.cell(r,15).value
    FreqTotal += Freq
    EarlyOrLate = shtPool.cell(r,16).value    
    Eq = Area + "\\" + pv + "\\" + eq
    if Area == "Topside":
        if hole == "FB":
            key =  Eq + "\\" + hole + "R_TLV\\Category " + weather
        else:
            key = Eq + "\\" + hole + "_TLV\\Category " + weather   
    # key = Eq + "\\" + hole + "\\Category " + weather
    
    
    
    Dxx.append([module,Freq, D00, D05, D10, D30, D60])    
    x = X[Eq]
    y = Y[Eq]
    h = ReleaseHeight[Eq]

    if ((EarlyOrLate == "EP") or (EarlyOrLate == "LP")):
        if EarlyOrLate == "EP":
            if key in EPSEP.keys():
                sep = EPSEP[key]
                pooltype = "EP"
                print("{:6s}|{:30s}|{:16s}|{:6.2f}|{:6.2f}|{:6.2f}|{:6s}|{:4s}|{:6s}|{:.2e}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}| {:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}| {:8.2f}".\
                    format(area,   pv,      eq + "\\"+hole+"\\"+weather+"-" + pooltype, x, y, h, module, hole, weather, Freq, sep, D00, D05, D10, D30, D60, H00,H05,H10,H30,H60))    
            else:
                print(key +" Not in EPSEP list")
        elif EarlyOrLate == "LP":
            if key in LPSEP.keys():
                sep = LPSEP[key]
                print("{:6s}|{:30s}|{:16s}|{:6.2f}|{:6.2f}|{:6.2f}|{:6s}|{:4s}|{:6s}|{:.2e}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}| {:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}|{:8.2f}| {:8.2f}".\
                    format(area,   pv,      eq + "\\"+hole+"\\"+weather+"-" + pooltype, x, y, h, module, hole, weather, Freq, sep, D00, D05, D10, D30, D60, H00,H05,H10,H30,H60))    
                pooltype = "LP"
            else:
                print(key +" Not in LPSEP list")
    else:
        # sep = 0.0
        # pooltype = "NA"
        print(key +" Neither EP or LP!")

    
    

    


