from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill

import matplotlib.pylab as pltparam

pltparam.rcParams["figure.figsize"] = (8,3)
pltparam.rcParams['lines.linewidth'] = 2
# pltparam.rcParams['lines.color'] = 'r'
pltparam.rcParams['axes.grid'] = True 


with open('v08_element.bin','rb') as v08dump:
    lEvent = dill.load(v08dump)

iExlFilename='v08_9mi'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    folder  = shPV.cell(r,3).value
    pv  = shPV.cell(r,15).value
    key = study  + "\\" +  folder  + "\\" + pv
    X[key] = shPV.cell(r,152).value
    Y[key] = shPV.cell(r,153).value
    r += 1
#Read X&Y and Put into lEvent
for e in lEvent:
    e.X = X[e.Study_Folder_Pv]
    e.Y = Y[e.Study_Folder_Pv]

#construct list of PV's
pvloc = []
listPV = X.keys()
for pv in listPV:
    if pvloc == []:
        pvloc = [X[pv],Y[pv]]
    else:
        pvloc = np.vstack([pvloc,[X[pv],Y[pv]]])
##nd of preparations for controus drawing


Resolution = 2
KK = np.arange(0,5) #Index for each time threshold to be read for each jet length
XX = np.linspace(99,224,50)
YY = np.linspace(-27,27,21)
TimeIndices = ['00','05','10','30','60']
KK = [1, 2, 3, 4]

PlotWhat = 'JetLength'
PlotWhat = 'HeatDose'
PlotWhat = 'SEP'

SEP = np.zeros([len(XX),len(YY),len(KK)])
for k in KK:
    ttl = TimeIndices[k]
    for i in range(0,len(XX)):
        for j in range(0,len(YY)):
            li = 0
            sep = np.zeros((len(lEvent),2))
            for e in lEvent:
                #read frequency
                sep[li,1] = e.JetFire.Frequency
                #check if the receptor is in the flame
                dx = XX[i]-e.X
                dy = YY[j]-e.Y
                rr = (dx*dx+dy*dy)
                jl = e.JetFire.JetLengths[k]
                if rr  <= (jl*jl):
                    #if the receptor is in the flame
                    sep[li,0] = e.JetFire.SEP
                else:
                    sep[li,0] = e.JetFire.SEP/(math.sqrt(rr)-jl)**2
                li += 1            
            # Sort 'sep' on the SEP
            # Read 'setp' corresponding to 1E-4
            # sep.sort(column=1)
            # sep_sorted = np.sort(sep,axis=1)[::-1] #sort by SEP in the descending order
            sep_sorted = sep[sep[:,0].argsort()][::-1]
            f_max2min = sep_sorted[:,1] #extract frequency
            cf = np.cumsum(f_max2min)
            # sep_sorted_with_cf = sep_sorted
            # sep_sorted_with_cf[:,1] = cf
            #interpolate on 'Column 2' and read from 'Column 1'
            if cf[-1] > 1E-4:
                fl = interp1d(cf,sep_sorted[:,0],kind='linear')
                sep_dim = fl(1E-4)
            else:
                print("Not dimensioning at ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[-1]))
                sep_dim = 0.            
            SEP[i,j,k-1] = sep_dim

    f = SEP[:,:,k-1].transpose()
    levels = [100, 200, 350]

    fig,ax = plt.subplots()

    img = plt.imread("RubyFPSODeckA.jpg")
    ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])

    cs1=ax.contourf(XX,YY,f,levels,colors=['yellow','magenta'],alpha=0.7)
    # cs1.cmap.set_under('white')
    # cs1.cmap.set_over('red')
    cs2=ax.contour(XX,YY,f,levels,colors=('k',),linewidths=(3,))
    ax.clabel(cs2,fmt='%3d',colors='k',fontsize=14)
    ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
    ax.set_aspect('equal')
    ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
    ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
    ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
    ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))

    # ax.xaxis.set_major_locator(plt.FixedLocator(XFramesAlleys))
    # ax.xaxis.set_major_formatter(plt.FixedFormatter(TickFrameAlleys))
    ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
    plt.title(ttl)
    # fig.savefig(fn+'v02.png')
    plt.show()



#Contours for Heat Dose
#Meanlingless to check all time threholds

HD = np.zeros([len(XX),len(YY)])
for i in range(0,len(XX)):
    for j in range(0,len(YY)):
        li = 0
        heat_dose = np.zeros((len(lEvent),2))
        for e in lEvent:
            #read frequency
            heat_dose[li,1] = e.JetFire.Frequency
            #check if the receptor is in the flame
            dx = XX[i]-e.X
            dy = YY[j]-e.Y
            rr = (dx*dx+dy*dy)

            jl = e.JetFire.JetLengths[1]
            if  jl == 0:
                #The fire will not last longer than 5 minutes
                jl = e.JetFire.JetLengths[0]
                duration = e.Discharge.Duration #in sec
                if rr  <= (jl*jl):
                    #if the receptor is in the flame
                    heat_dose[li,0] += e.JetFire.SEP*duration
                else:
                    heat_dose[li,0] += e.JetFire.SEP/(math.sqrt(rr)-jl)**2*duration
                # heat_dose[li,0] = 0 #to ignore transient(?) jet fire shorter than 5 minutes
            else:
                for k in KK:
                    jl = e.JetFire.JetLengths[k-1]
                    if jl < 40:
                        sep = jl*8.75 #e.JetFire.SEP as a fitting to SEP vs jet length
                    else:
                        sep = e.JetFire.SEP #should be 350 or so
                    duration = e.Discharge.Ts[k] - e.Discharge.Ts[k-1] #in sec. if k=1, duration = 5 min, 300 sec
                    if rr  <= (jl*jl):
                        #if the receptor is in the flame
                        heat_dose[li,0] += sep*duration
                    else:
                        heat_dose[li,0] += sep/(math.sqrt(rr)-jl)**2*duration
            li += 1            
        # Sort 'sep' on the SEP
        # Read 'setp' corresponding to 1E-4
        # sep.sort(column=1)
        # sep_sorted = np.sort(sep,axis=1)[::-1] #sort by SEP in the descending order
        heat_dose_sorted = heat_dose[heat_dose[:,0].argsort()][::-1]
        f_max2min = heat_dose_sorted[:,1] #extract frequency
        cf = np.cumsum(f_max2min)
        # sep_sorted_with_cf = sep_sorted
        # sep_sorted_with_cf[:,1] = cf
        #interpolate on 'Column 2' and read from 'Column 1'
        if cf[-1] > 1E-4:
            fl = interp1d(cf,heat_dose_sorted[:,0],kind='linear')
            heat_dose_dim = fl(1E-4)
        else:
            print("Not dimensioning at ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[-1]))
            heat_dose_dim = 0.            
        HD[i,j] = heat_dose_dim

contour_levels = [30000,  60000,  105000, 120000,   180000,   210000,   360000,  630000,   720000, 1260000]
contour_colors = ['green','lime',  'cyan', 'yellow','olive',  'magenta', 'purple','crimson', 'red']



ttl = 'HeatDose'
f = HD.transpose()
#Levels for Heat Dose
# 100 x 5min
# 200 x 5 min
# 350 x 5 min = where Durations are given in seconds
# 100 x 10 min = 200 x 5min 60000
# 200 x 10min =   120000
# 350 x 10 min =  210000
# 100 x 30 min =  180000
# 200 x 30 min =  360000
# 350 x 30 min =  630000
# 200 x 60 min =  720000
# 350 x 60 min = 1260000

#https://stackoverflow.com/questions/12274529/how-to-smooth-matplotlib-contour-plot
# Resample your data grid by a factor of 3 using cubic spline interpolation
# f_resampled = scipy.ndimage.zoom(f,3)
sigma = 0.7
f_resampled = gaussian_filter(f,sigma)


fig,ax = plt.subplots(figsize=(8,3))

img = plt.imread("RubyFPSODeckA.jpg")
ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])

# cs1=ax.contourf(XX,YY,f,levels,colors=['yellow','magenta'],alpha=0.7)
cs1=ax.contourf(XX,YY,f_resampled,contour_levels,colors=contour_colors,alpha=0.7)
fig.colorbar(cs1)
cs1.cmap.set_under('w')
cs1.cmap.set_over('k')
cs2=ax.contour(XX,YY,f_resampled,contour_levels,colors=('k',),linewidths=(1,))
ax.clabel(cs2,fmt='%3d',colors='k',fontsize=10)
ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
ax.set_aspect('equal')
ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))

ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
ax.yaxis.grid(b=True, linestyle='--',linewidth=2)
plt.title(ttl)
# fig.savefig(fn+'v02.png')
plt.show()






#Contour for jet fire impingement
F = np.zeros([len(XX),len(YY),len(KK)])
#For each time threshold
for k in KK:
    for e in lEvent:
    #Evaluat rectangle 
        jl = e.JetFire.JetLengths[k]
        x = e.X
        y = e.Y
        #Add frequency to point in the rectangle [x-j,x+j] [ y-j,y+j]
        xi = np.argwhere((XX >= x-jl) & (XX <= x+jl))
        yi = np.argwhere((YY >= y-jl) & (YY <= y+jl))
        for j in yi:
            for i in xi:
                dx = XX[i]-e.X
                dy = YY[j]-e.Y
                if ((dx*dx+dy*dy) < (jl*jl)):
                    F[i,j,k] += e.JetFire.Frequency
        


        
XXX, YYY = np.meshgrid(XX,YY)
k=4
ttl = TimeIndices[k]
f = F[:,:,k].transpose()
# ll = size(XXX)

fig,ax = plt.subplots()

img = plt.imread("RubyFPSODeckA.jpg")
ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])


levels = [1E-6, 1E-5, 1E-4,1E-3]
cs1=ax.contourf(XX,YY,f,levels,colors=['green','yellow','magenta'],alpha=0.7)
# cs1.cmap.set_under('white')
# cs1.cmap.set_over('red')
cs2=ax.contour(XX,YY,f,levels,colors=('k',),linewidths=(3,))
ax.clabel(cs2,fmt='%.1e',colors='k',fontsize=14)
ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
ax.set_aspect('equal')
ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))

# ax.xaxis.set_major_locator(plt.FixedLocator(XFramesAlleys))
# ax.xaxis.set_major_formatter(plt.FixedFormatter(TickFrameAlleys))
ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
plt.title(ttl)
# fig.savefig(fn+'v02.png')
plt.show()