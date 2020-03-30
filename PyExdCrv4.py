import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import numpy as np
import dill
import unicodedata
import re


def print_a_event(kwd,lEvent):
    i=0
    for e in lEvent:
        if kwd in e.Key:
            print(i,e)
        i += 1

def slugify(value):
    """
    Normalizes string, converts to lowercase, removes non-alpha characters,
    and converts spaces to hyphens.
    """
    # import unicodedata
    # value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore')
    # value = unicode(re.sub('[^\w\s-]', '', value).strip().lower())
    # value = unicode(re.sub('[-\s]+', '-', value))
    value = re.sub('[^\w\s-]', '', value).strip().lower()
    value = re.sub('[-\s]+', '-', value)
    return value

NumDirections = 6 # 6 cones to cover the all release direction for a sphere


TVDPLot = True


#Topside
Area = 'ProcessArea'
iExlFilename='Bv06_i'
cExlFilename='Bv06_c'
# cExlFilename='Bv05jalLM_Hull_c'
# SysLen = 7 #Length of 'v08_9m' + 1
# SysLen = 23 #Length of 'RUBY_FRA_Rev.B_v01_shk' + 1
# SysLen = len('RUBY_FRA_Rev.B_v01_shk') + 1
EqvsModule = {'020-01': 'S05', '020-02':'S05', \
    '020-03':'S04', '020-04':'S04', \
        '021-01':'S02','023-01':'P03','023-02':'P03', \
            '023-03':'S03', '023-04':'S03', \
                '024-01':'P04', '024-02':'P04','024-03':'P04', '025-01':'P04', '025-02':'P04', \
                    '025-03':'S03','027-01':'P05', '027-02':'P05', '027-03':'P05', \
                        '045-01':'P03','045-02':'P03','046-01':'P02','013-01':'S05'}

SFXFiles = ['013-01-C1-G', '013-01-C2-G', '013-01-N1-G', '013-01-N2-G', \
    '020-01-01-G', '020-01-02-L', '020-02-01-Li', '020-02-02-G', '020-02-03-Lo', '020-03-01-Li', '020-03-02-G', '020-03-03-Lo', '020-04-01-L', '020-04-02-G', '020-05-01-G', '020-05-02-L', \
    '023-01-01-G-DA', '023-01-01-G-DB', '023-01-02-L', '023-02-01-G-DA', '023-02-01-G-DB', '023-02-02-L', '023-03-01-G-DA', '023-03-01-G-DB', '023-03-01-G-DC', '023-03-02-L','023-03-03-L', '023-04-G',\
    '024-01-01-G','024-01-02-L', '024-02-G-DA', '024-02-G-DB', '024-02-G-DC', '024-03-01-G', '024-03-02-L',  \
    '025-01-01-G', '025-01-02-L', '025-02-01-G', '025-02-02-L', '027-01-G-DA', '027-01-G-DB', '027-01-G-DC', '027-02-G', '027-03-G',  \
    '043-03-G', '045-01-G', '045-02-01-G', '045-02-02-L', '045-03-G', '046-02-L', '046-03-L',\
    '021-01-01-L','046-01-01-L','046-01-02-L','046-01-03-L','046-01-04-L','062-01-01-L','043-01-01-G','043-01-02-L','043-02-01-G','043-02-02-L']
TVDprefix = ''
element_dump_filename = 'Bv06_dump'
#Topside Parameters End

#Offloaidng ara analysis
# 'Flammable dispersion' is not relevant??? All vaporized??
cExlFilename='Bv06_O_c'
SFXFiles = ['021-02-L']
element_dump_filename = 'Bv06_offloading_dump'

#Utlity analysis
cExlFilename='Bv06_u_c'
SFXFiles = ['045-04-G','062-01-02-L']
element_dump_filename = 'Bv06_utility_dump'

#Hull deck analysis
cExlFilename='Bv06_h_c'
SFXFiles = ['131-01-L-fwdP','131-01-L-fwdS','131-01-L-aftP','131-01-L-aftS','131-02-L-fwdP','131-02-L-fwdS','131-02-L-aftP','131-02-L-aftS','057-01-02-L','057-01-01-G','046-04-L','021-01-02-L','013-06-L','013-05-L']
element_dump_filename = 'Bv06_hull_dump'



iExl=load_workbook(filename=iExlFilename+'.xlsx')
cExl=load_workbook(filename=cExlFilename+'.xlsx')
fExl=load_workbook(filename='Leak_Freq_Hole_Size.xlsx')
shFreq = fExl['RESULTS - Hole size']


shPV = iExl['Pressure vessel']
shtTLV = iExl['Time varying leak']
shDc = cExl['Discharge']
shJet = cExl['Jet fire']
shDispersion = cExl['Flammable Dispersion']
shExplosion = cExl['Explosions']


def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    # if m>5:
    #     jf = -13.2+54.3*math.log10(m)
    # elif m>0.1:
    #     jf= 3.736*m + 6.
    # else:
    #     jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
def mfit(jl):
    m = np.power(10,math.log10(jl/2.8893)/0.3728) / 55.5
    return m

class Event:
    def __init__(self,study_folder_pv,scenario,weather,dc=None,jf=None,epf=None,lpf=None,exp=None,freq=None,x=None,y=None,rh = None, \
        hole = None,holemm = None, PP = None, TT = None,fireball = None, dispersion = None, esd = True, bdv = True,module = None,deck= None,pesd = 1, pbdv = 1):        
        self.Study_Folder_Pv = study_folder_pv
        self.Path = study_folder_pv + "\\" + scenario
        self.Weather = weather
        study,pv, folder = study_folder_pv.split("\\")
        self.Key = pv + "\\" + scenario + "\\" + weather
        self.Frequency = freq
        self.Discharge = dc
        self.JetFire = jf
        self.EarlyPoolFire = epf
        self.LatePoolFire = lpf
        self.Explosion = exp
        self.Fireball = fireball
        self.Dispersion = dispersion
        self.TVD = []
        self.TVDRead = False
        self.Module = module
        self.Deck = deck
        self.X = x
        self.Y = y
        self.ReleaseHeight = rh #should be 'Elevation.' Or Elevation + ReleaseHeight?
        self.Hole = hole
        self.Holemm = holemm
        self.Pressure = PP
        self.Temperature = TT        
        self.PImdIgn = 1
        self.PDelIgn = 0
        self.PExp_Ign = 0.21 #WOAD as referred by by IP & UKOOA as the 'probability of explosion given ignition'
        self.ESDSuccess = esd
        self.BDVSuccess = bdv
        self.BDVDia = 0
        self.jfscale = 1      
        self.PESD = pesd
        self.PBDV = pbdv
        """     def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # RelDuration = self.TVD[-1,0]
            a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:10s} | {:40s} | {:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} | {:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}*".\
                format(ts,   folder,      pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], self.JetFire.Frequency, self.JetFire.SEP, a, b, c, d,e)
        else:
            fmt = "{:10s} | {:40s} | {:20s} | {:20s} | {:20s} | {:20s} | {:.2e} | {:8.2f} | {:8.2f}".\
                format(ts,   folder,      pv, self.Module, scenario, self.Weather, self.Frequency, self.Discharge.ReleaseRate,self.Discharge.Duration)
        
        return fmt """
    def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # print(self.Path,self.Dispersion,self.JetFire,self.EarlyPoolFire,self.LatePoolFire,self.Explosion,self.Fireball)
            # RelDuration = self.TVD[-1,0]
            # a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} \
                | {:8.2f} [barg]| {:8.2f} [degC]| | ESD : {:}| BDV  {:} | Freq: {:8.2e}\n".\
                format(pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], \
                self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess, self.Frequency)
            
            if self.Dispersion != None:
                fmt += "\tDispersion: Frequency {:8.2e}, Distance to LFL {:}\n".format(self.Dispersion.Frequency,self.Dispersion.DfMax)
            else:
                fmt += "\tDispersion: No dispersion\n"
            
            if self.JetFire != None:
                fmt += "\tJF: Frequency {:8.2e}, Length {:6.1f}\n".format(self.JetFire.Frequency,self.JetFire.Length)
            else:
                fmt += "\tJF: No Jet Fire\n"
            
            if self.Explosion != None:
                fmt += "\tExplosion: Frequency {:8.2e}, Overpressure {:6.1f} to {:6.1f} [m]\n".format(self.Explosion.Frequency,self.Explosion.Overpressures[-1],self.Explosion.Distance2OPs[-1])
            else:
                fmt += "\tExplosion: No Explosion\n"
            
            if self.EarlyPoolFire != None:
                fmt += "\tEPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.EarlyPoolFire.Frequency,self.EarlyPoolFire.Diameter)
            else:
                fmt += "\tEPF: No Early Pool Fire\n"
            if self.LatePoolFire != None:
                fmt += "\tLPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.LatePoolFire.Frequency,self.LatePoolFire.Diameter)
            else:
                fmt += "\tLPF: No Late Pool Fire\n"
            if self.Fireball != None:
                fmt += "\tFireball: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.Fireball.Frequency,self.Fireball.Diameter)
            else:
                fmt += "\tFireball: No Fireball\n"
        
        else:
            ts,pv,IS,hole = self.Path.split("\\")
            f = "{:15s} {:4s} ({:6.1f},{:6.1f},{:6.1f}) {:6.1f} {:6s} P({:8.2f}) T({:8.2f}) ESD({:}) BDV({:}) | Freq: {:8.2e}".\
                        format(IS, self.Module, self.X, self.Y,  self.ReleaseHeight, self.Holemm, self.Weather,self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess,self.Frequency)
            if self.Dispersion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.JetFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.EarlyPoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.LatePoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Explosion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Fireball != None:
                f += "  O  "
            else:
                f += "  X  "            
            fmt = f
        
        return fmt
        # return self.Key
    def EventTree(self):
        # Pwss = 1/8*1/2 #8 wind directions & 2 wind velocities
        Pwss = 1
        Pi = self.PImdIgn # Probability of immediate ignition #input into SAFETI hole        
        Pdi = self.PDelIgn

        Ph = 4/6 #horizontal fraction for vertical and horizontal release cases
        Pj_h = 1/4*Ph # probability of horizontal jet fire Pjh x ph = 4/6 x 1/4 = 1/6 with no accompanying pool fire
        Pp_h = 1/2*Ph #probability of pool fire upon horizontal release with no accompanying jet fire
        Pjp_h =0.5/4*Ph #probability of jet & pool fire upon horizontal release

        Pj_v = 1/2*(1-Ph)#probabilty of vertical jet fire upon vertical release
        Pp_v = 1/2*(1-Ph)#
        Pjp_v = 0.5/4*(1-Ph)#

        #short release
        Ps = 1 #fraction for short release effects; release is shorter than cut-off time '20 s'
        Pbp = 1 #prob of bleve & pool fire
        Pb = 0 #prob of bleve without a pool fire
        Pfp = 0 #prob for short duration release immediate ignition resulting in flash fire and pool
        Pf = 0 #Flash fire along
        Pp = 0#prob for short duration release immediate ignition resulting in  to have pool fire
        Pep = 0#prob for short duration release immediate ignition resulting in  to have immediate explosion with pool fire
        Pe = 0#prob for short duration release immediate ignition resulting in  to have immediate explosion without pool fire
        Pp = 0#prob for short duration release immediate ignition resulting in  to have early pool fire without explosion


        Prp = 0.15 #prob of residual pool fire
        Po = (1-Pi)*Pwss*Pipr*Prp#prob of igniton of residual pool
        Pf_d = 0.6 #prob of delayed flash fire
        Pe_d = 0.4 #prob of explosion upon delayed ignition

        #Immediate, Not short
        P_i_ns_h_j = (1-Ps)*Ph*Pwss*Pj_h*Pi
        P_i_ns_h_p = (1-Ps)*Ph*Pwss*Pp_h*Pi
        P_i_ns_h_jp = (1-Ps)*Ph*Pwss*Pjp_h*Pi
        P_i_ns_v_j = (1-Ps)*(1-Ph)*Pwss*Pj_v*Pi
        P_i_ns_v_p = (1-Ps)*(1-Ph)*Pwss*Pp_v*Pi
        P_i_ns_v_jp = (1-Ps)*(1-Ph)*Pwss*Pjp_v*Pi
        #Immediate, Short
        P_i_s_BP = Ps*Pwss*Pbp*Pi #BLEVE & Pool
        P_i_s_B = Ps*Pwss*Pb*Pi #BLEVE only
        P_i_s_FP = Ps*Pwss*Pfp*Pi #Flash fire and pool
        P_i_s_F = Ps*Pwss*Pf*Pi #Flash fire only
        P_i_s_EP = Ps*Pwss*Pep*Pi #Explosion & pool fire
        P_i_s_E = Ps*Pwss*Pe*Pi #Explosion only
        P_i_s_P = Ps*Pwss*Pp*Pi #Pool fire only

        #Delayed
        P_d_FP = (1-Pi)*Pwss*Pfp*Pdi #Delayed flash and pool fire
        P_d_E = (1-Pi)*Pwss*Pdd*Pdi #Delayed explosion
        P_d_p = (1-Pi)*Pwss*Prp*Pirp #Redisual pool fire

        print("----Imd Ignition---Not short--|-Horizontal---Jet fire       : {:8.2e}".format(P_i_ns_h_j))
        print("                 |             |             |-Pool fire      : {:8.2e}".format(P_i_ns_h_p))
        print("                 |             |             -Jet & Pool fire: {:8.2e}".format(P_i_ns_h_jp))
        print("                 |             |-Vertical-----Jet fire       : {:8.2e}".format(P_i_ns_v_j))
        print("                 |                           |-Pool fire      : {:8.2e}".format(P_i_ns_v_p))
        print("                 |                           -Jet & Pool fire: {:8.2e}".format(P_i_ns_v_jp))
        print("                 --Short----------------------BLEVE & Poolfire: {:8.2e}".format(P_i_s_BP))
        print("                                            |-BLEVE only      : {:8.2e}".format(P_i_s_B))
        print("                                            |-Flash & Pool fire:{:8.2e}".format(P_i_s_FP))
        print("                                            |-Flash only: {:8.2e}".format(P_i_s_F))
        print("----Delayed-------Ignited--------------------Flash and Pool fire:{:8.2e}".format(P_d_FP))
        print("                                            |- Explosion : {:8.2ef}".format(P_d_E))
        print("                                            |- Pool fire: {:8.2e}".format(P_d_p))
        

class Discharge:
    def __init__(self,rr,duration,liqfrac = 0.):
        self.ReleaseRate = rr
        self.Duration = duration
        self.LiquidMassFraction = liqfrac
        self.Ts = [0.,0.,0.,0.,0.]
        self.Ms = [0.,0.,0.,0.,0.]
        self.RRs = [0.,0.,0.,0.,0.]        

class Dispersion:
    def __init__(self,fdisp,dfmax,dfmin,wf,dmax,dmin,w,ffmax,ffw):
        self.Frequency = fdisp        
        
        # self.Distance2LFLFraction = d2f
        # self.Distance2LFL = d
        
        #Column P: max d, Q: min d, R: max w for LFL fraction
        #Column S: max d, T: min d, U: max w for LFL fraction
        self.DfMax = dfmax
        self.DfMin = dfmin
        self.Wf = wf
        self.DMax = dmax
        self.DMin = dmin
        self.W = w

        self.FFMaxDistance = ffmax
        self.FFWidth = ffw

    def __str__(self):
        
        fmt = "{:.2e} | {:8.2f}m to LFL Fraction | {:8.2f}m to LFL".\
            format(self.Frequency, self.DfMax, self.DMax)        
        return fmt

class JetFire:
    def __init__(self,length,sep,jff):
        self.Length = length
        self.SEP = sep
        self.Frequency = jff
        self.Ts = [0.,0.,0.,0.,0.]
        self.JetLengths = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.JetLengths
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)        
        return fmt

class EarlyPoolFire:
    def __init__(self,diameter,sep,epff):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

class LatePoolFire:
    def __init__(self,diameter,sep,epff):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

class Explosion:
    def __init__(self,fexp,op,d2op):
        self.Overpressures = op
        self.Distance2OPs = d2op
        self.Frequency = fexp        
    def __str__(self):
        a,b,c = self.Distance2OPs
        pa,pb,pc = self.Overpressures
        fmt = "{:.2e} | {:8.2f}barg to {:8.2f} | {:8.2f}bar to {:8.2f} | {:8.2f} barg to {:8.2f}".\
            format(self.Frequency, pa, a, pb, b, pc,c)        
        return fmt

class Fireball:
    def __init__(self,ffb,D,sep):
        self.Diameter = D
        self.SEP = sep
        self.Frequency = ffb       
    def __str__(self):
        
        fmt = "{:.2e} | Diameter: {:8.2f} | SEP {:8.2f}".\
            format(self.Frequency, self.Diameter, self.SEP)        
        return fmt

#Read Leaks
#listLeak = []
ReleaseHeight= {} #from vessel bottom, Column 'V'? Shall be elevation Columnt 'T'?
Frequency ={}
PImdIgn ={}
PDelIgn = {}
PExp ={}
ReleaseRate = {}
Duration = {}
J00 = {}
J05 = {}
J10 = {}
J30 = {}
J60 = {}
JetSEP = {}
JetFrequency = {}
Tiso = {}
Tbdv = {}
Hole = {}
Holemm = {}
PP = {} #Pressure of the PV
TT  = {} #Temp of the PV
ESDSuccess ={}
BDVSuccess ={}
BDVDia ={}
X = {}
Y = {}
Z = {}
Material = {}



iIS=load_workbook(filename='IS_v12_shk.xlsx')
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 82:
    pv = shIS.cell(r,5).value
    IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    Modules[pv] = shIS.cell(r,7).value
    Deck[pv] =  shIS.cell(r,8).value
    if shIS.cell(r,24).value != None:
       nedvs = shIS.cell(r,24).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1

r=5
while shFreq.cell(r,2).value == "Full pressure leak":     
    pv  = shFreq.cell(r,1).value 
    Frequency[pv] = [0.,0.,0.,0.]
    if pv[-1] == "G":
        Frequency[pv][0] = shFreq.cell(r,4).value #Gas SM
        Frequency[pv][1] = shFreq.cell(r,5).value # ME
        Frequency[pv][2] = shFreq.cell(r,6).value # MA
        Frequency[pv][3] = shFreq.cell(r,7).value # LA
    else:
        Frequency[pv][0] = shFreq.cell(r,10).value #Liq SM
        Frequency[pv][1] = shFreq.cell(r,11).value # ME
        Frequency[pv][2] = shFreq.cell(r,12).value # MA
        Frequency[pv][3] = shFreq.cell(r,13).value # LA
    r += 3
numIS = (r-5)/3



#Read 'Presure vessel' from Input
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv
    Material[key] = shPV.cell(r,9).value
    TT[key] = shPV.cell(r,16).value
    PP[key] = shPV.cell(r,16).value
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    r += 1
numPV = r-63




#Read 'Time varying leak' from Input
#for r in range(6,numscn):
r = 63
while shtTLV.cell(r,1).value == "Yes":
    study = shtTLV.cell(r,2).value    
    relheight = shtTLV.cell(r,32).value
    equip  = shtTLV.cell(r,8).value
    folder = shtTLV.cell(r,9).value
    hole = shtTLV.cell(r,20).value 
    holemm = shtTLV.cell(r,21).value
    relheight = shtTLV.cell(r,25).value
    # key = study + "\\" + equip + "\\" + folder + "\\"  + hole
    key_hole = equip+"-"+hole
    ReleaseHeight[equip] = relheight

    #Construci disctionarys with keys as 'Path'
    Hole[key_hole] = hole
    Holemm[key_hole] = holemm
    #SAFETI
    # Frequency[key]  = shtTLV.cell(r,38).value
    # PImdIgn[key] = shtTLV.cell(r,44).value
    # PDelIgn[key] = shtTLV.cell(r,46).value
    # PExp[key] = shtTLV.cell(r,48).value
    #PHAST doesn't provide this frequency & probability    
    
    # PHAST
    ESDSuccess[key_hole] = shtTLV.cell(r,44).value
    Tiso[key_hole] = shtTLV.cell(r,45).value
    BDVSuccess[key_hole] = shtTLV.cell(r,47).value
    Tbdv[key_hole] = shtTLV.cell(r,48).value
    BDVDia[key_hole] = shtTLV.cell(r,49).value
    r=r+1
    print("Reading Time varying leaks: ",key_hole)
numLeak = r-63



lEvent = [] #List of class Event
#Construct Event list
#Read Discharge, shDc
# And evaluate ignition probability'
# weathers = ['1.2/F', '6.6/D']
weathers = ['2.9F','7.7D','14.5D']
HoleSizes = [7.1, 36.1, 111.8, 150]
dk=0
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shDc.max_row+1):
    study_folder_pv = shDc.cell(r,1).value
    hole = shDc.cell(r,2).value
    if hole != "Catastrophic rupture":
        study,pv,folder = study_folder_pv.split("\\")
        scenario = shDc.cell(r,2).value
        weather = shDc.cell(r,3).value
        path = study_folder_pv + "\\" + scenario
        # key = study_folder_pv + "\\" + scenario + "\\" + weather
        key_hole = pv +"-"+ hole
        key_hole_weather = pv +"-"+hole+"-"+weather
        liqmassfrac = shDc.cell(r,14).value
        if weather in weathers:
            rr = shDc.cell(r,10).value
            ReleaseRate[key_hole] = rr
            duration = shDc.cell(r,11).value
            x = X[pv]
            y = Y[pv]
            z = Z[pv]
            Duration[key_hole] = duration
            basefreq = 0.
            #Frequency allocatin
            if hole[0:2] in ['SM','ME','MA','LA']:
                holesizeindex = ['SM','ME','MA','LA'].index(hole[0:2])
                if pv in Frequency.keys():                
                    basefreq = Frequency[pv][holesizeindex]
                else:
                    for k in Frequency.keys():
                        if k in pv:
                            fkey = k
                    basefreq = Frequency[fkey][holesizeindex]
            elif hole[0:2] == "LM":
                for k in range(0,4):
                    if Holemm[key_hole] > HoleSizes[k]:
                        continue
                    else:
                        break
                holesizeindex = k                
                if pv in Frequency.keys():
                    for k in range(holesizeindex,4): #Add all frequency for larger hole size categories for LM
                        basefreq += Frequency[pv][k]
                        Frequency[pv][k] = 0.
                    Frequency[pv][holesizeindex] = basefreq                    
                else:
                    for k in Frequency.keys():
                        if k in pv:
                            fkey = k
                    for k in range(holesizeindex,4):
                        basefreq += Frequency[fkey][k]
                        Frequency[fkey][k] = 0.
                    Frequency[fkey][holesizeindex] = basefreq
            
            leak_freq = IS_sub[pv]*basefreq#actually Frequency[pv] = 1            

            
            # aEvent = Event(study_folder_pv,scenario,weather,freq=Frequency[path],dc=Discharge(rr,duration),hole=Hole[path])
            p_ESDfail = 0.01*numESDVs[pv]
            if hole[4]=="O":
                esd = True
                p_esd_branch = 1-p_ESDfail
                leak_freq *= 1-p_ESDfail
            elif hole[4] == "X":
                esd = False
                p_esd_branch = p_ESDfail
                leak_freq *= p_ESDfail
            else:
                print(pv,hole,'ESDV value wrong')
                break
            #BDV failure into Event Frequency
            if hole[-1]=="O":
                bdv = True
                p_bdv_branch = (1 - 0.005)
                leak_freq *= (1 - 0.005)
            elif hole[-1]=="N":
                bdv = "None"
                p_bdv_branch = 1
            elif hole[-1]=="X":
                bdv = False
                p_bdv_branch = 0.005
                leak_freq *= 0.005
            else:
                print(pv,hole,'BDV value wrong')
                break

            if weather == '2.9F':
                leak_freq *= 0.486
            elif weather == '7.7D':
                leak_freq *= 0.471
            elif weather == '14.5D':
                leak_freq *= 0.044
            else:
                print('Wrong wetaher', key_hole_weather)
                break



            aEvent = Event(study_folder_pv,hole,weather,freq=leak_freq,dc=Discharge(rr,duration,liqfrac=liqmassfrac),hole=Hole[key_hole],holemm = Holemm[key_hole],\
                esd=esd,bdv=bdv,x=x,y=y,rh=z,PP=PP[pv],TT=TT[pv],module=Modules[pv],deck=Deck[pv],pesd=p_esd_branch, pbdv=p_bdv_branch)
            lEvent.append(aEvent)

        else:
            print("{} Discharge events read".format(dk))
            break
        dk=dk+1

#End of constructing list of Events
#Up to now, read IS, '_i', '_c[Discharge]'

IgnProb = {}
IgnProb['023-01']=[0.00050007132890309,0.00120241794400881,0.0121099047419176,0.0133941985161508]
IgnProb['023-02']=[0.000500009730687852,0.00112728346839808,0.0114299096696923,0.0124894860472479]
IgnProb['045-01']=[0.000500092602037187,0.00214457708972392,0.0159824850536888,0.015707932976315]
IgnProb['024-01']=[0.000500049238017383,0.00135755735655058,0.0127160815903571,0.0137890702655233]
IgnProb['024-02']=[0.000500092602037187,0.00177918758871906,0.014582259955286,0.0152682624925632]
IgnProb['024-03']=[0.000500092602037187,0.00187273284139629,0.0148864999002958,0.0153843297225681]
IgnProb['025-01']=[0.000500066037325877,0.00155843931004843,0.0138293983168883,0.0148186113336855]
IgnProb['025-02']=[0.000500076839107057,0.001583080559352,0.0138624647673917,0.0148213944403002]
IgnProb['027-01']=[0.000500092602037187,0.00141234355179696,0.0130284920753651,0.014098525976969]
IgnProb['027-02']=[0.000500092602037187,0.00167828753177232,0.0139540201469402,0.0148124734230255]
IgnProb['023-03']=[0.000500074056674887,0.00134865578738282,0.0127446106070196,0.013756233693594]
IgnProb['023-04']=[0.000500014189107731,0.00114561328550342,0.0121656333278765,0.0136982452845986]
IgnProb['024-01']=[0.000500056109924782,0.00140883072048731,0.013214286058242,0.0148872691410471]
IgnProb['045-02']=[0.000500077778731805,0.00141477302560298,0.012839767974552,0.0139304583821647]
IgnProb['020-01']=[0.000500081597970954,0.00198507390262245,0.0159033776172752,0.0159730079240214]
IgnProb['020-02']=[0.000500044233038885,0.00127252532590165,0.0126086841965461,0.0139360824741458]
IgnProb['013-01']=[0.000500081597970954,0.00125827417928573,0.0124609025752969,0.0136146062289823]
IgnProb['020-03']=[0.000500037598733664,0.00117389329582008,0.0122180292562774,0.0135140045120202]
IgnProb['020-04']=[0.000500009985715746,0.00113070577630583,0.0120388416453966,0.0131069382140561]
IgnProb['027-03']=[0.000500087509063428,0.00137082995866607,0.0129208012390835,0.0142136300044506]
IgnProb['General']=[0.000502,0.001677,0.017112,0.017916]


IgnitionModel = "OLF"
p = []
#Set Ignition probability after reading the Discharge result
for e in lEvent:    
    # study,pv,pv_hole,hole,weather = e.Key.split("\\")
    pv,hole,weather = e.Key.split("\\")        
    key_hole = pv +"-"+ hole
    rr = ReleaseRate[key_hole]
    for k in IgnProb.keys():
        if k in pv:
            p = IgnProb[k]   
    
    if IgnitionModel == "OLF":
        # Adjustment for liquid mass fraction
        rr *= (1-min(e.Discharge.LiquidMassFraction,0.95))        
        if p == []:
            p = IgnProb['General']
        if rr < 1:
            e.PImdIgn = 0.0005
            e.PDelIgn = p[0] - e.PImdIgn
            # e.PDelIgn = p[0]
        elif rr < 10:
            e.PImdIgn = 0.001
            e.PDelIgn = p[1] - e.PImdIgn        
            # e.PDelIgn = p[1]
        elif r < 30.:
            e.PImdIgn = 0.01        
            e.PDelIgn = p[2] - e.PImdIgn        
            # e.PDelIgn = p[2]
        elif r >= 30.:        
            e.PImdIgn = 0.01
            e.PDelIgn = p[3] - e.PImdIgn        
            # e.PDelIgn = p[3]
        # Adjustment for liquid mass fraction
        # e.PImdIgn *= (1-min(e.Discharge.LiquidMassFraction,0.95))
        # e.PDelIgn *= (1-min(e.Discharge.LiquidMassFraction,0.95))
        # print(key_hole, rr, e.PImdIgn)
    elif IgnitionModel == "UKOOA": #No.22 Offshore FPSO Gas or Liquid
        if "-L" in pv:
            #26 FPSO Liquid
            max_p =0.028
            min_p = 0.001
            pointa = 0.1
            grad_a = 0.468558
            offset_a = -2.49002
            pointb = 100
            if rr < pointa:
                pign = np.power(10,grad_a*np.log1o(rr)+offset_a)
                if pign < min_p:
                    pign = min_p
            if rr >= pointb:
                pign = max_p
        else:
            #22 Offshor Process Gas Congested or Mechanical Vented Module
                        
            #24 Offshore FPSO Gas
            max_p =0.15
            min_p = 0.001
            pointa = 0.1
            grad_a = 0.072551
            offset_a = -2.88606
            pointb = 1
            grad_b = 1.213764
            offset_b = -2.88606
            pointc = 50

            if rr < pointa:
                pign = np.power(10,grad_a*np.log1o(rr)+offset_a)
                if pign < min_p:
                    pign = min_p
            elif rr < pointb:
                pign = np.power(10,grad_b*np.log1o(rr)+offset_b)
                if pign > max_p:
                    pign = max_p
            elif rr >= pointc:
                pign = max_p
    else:
        # rr *= (1-min(e.Discharge.LiquidMassFraction,0.95))    
        if p == None:
            p = IgnProb['General']
        if rr < 1:
            e.PImdIgn = p[0]
        elif rr < 10:
            e.PImdIgn = p[1]
        elif r < 30.:
            e.PImdIgn = p[2]
        elif r >= 30.:        
            e.PImdIgn = p[3]

        
    # e.PImdIgn = PImdIgn[path]
    # e.PDelIgn = PDelIgn[path]
    # e.PExp_Ign = PExp[path]
   



#Read TV Discharge data and append the info on each Event ... equip - hole - weather

for sfx in SFXFiles:
# for sfx in ['013-01-C2-G']:
# for sfx in ['021-01-01-L']:
    #sfx = '020-01'
    tvExl = load_workbook(filename=TVDprefix + sfx+'.xlsx')
    print("Reading ...", sfx)
    sh = tvExl['Sheet1']
    new_scn = False
    #iterte for all row
    #read total row

    r = 1
    while r < sh.max_row:
        rv = sh.cell(r,1).value    
        #read the path at the row next toe 'Sceanrio ...'
        if rv != None:
            if "Scenario" in rv:
                #print (r, "in the loop")
                #print(rv)            
                r = r + 1
                path = sh.cell(r,1).value
                filename,study,pv,folder,hole = path.split("\\")                
                # path = path[SysLen:]
                path = study +"\\"+pv +"\\"+folder +"\\"+hole

                r = r + 2
                weather = sh.cell(r,1).value
                weather = weather[9:]
                #scn = path[20:]+"-"+weather[9:]
                scn = path +"\\"+weather
                r = r + 3
                t = sh.cell(r,1).value # read t=0 data
                mm = sh.cell(r,2).value
                rr = sh.cell(r,3).value
                m0 = mm
                r0 = rr
                TVD = [t, mm, rr ]            
                
                while t != None:
                    mm = sh.cell(r,2).value
                    rr = sh.cell(r,3).value
                    TVD = np.vstack([TVD, [t, mm, rr]])  
                    # if (("Flow line" in path) and ("NE_TLV" in path) and (t==3600)):
                    #     print([t, mm, rr])
                    #     print(TVD[-1,:])
                    tp = t
                    r = r + 1
                    t = sh.cell(r,1).value    
                                
                # if tp==3600:
                    # print(path, weather, TVD[-1,:])
                #print("Scenario {} ends at time {:8.2f} \n with the remaining mass {:10.2f}( \% of the inventory) and Release rate: {:8.2f})".\
                #    format(scn, tp, mm/m0*100, rr))
                #print("End of scenario {}".format(path+weather))        
                #print(r, rv)                


                #print out release rate at t=0, t=5min, t=10min, t=30min, t = 60min
                T_gate = [0, 300, 600, 1800, 3600]
                M_gate = [TVD[0,1], 0., 0., 0., 0.]
                RR_gate = [TVD[0,2], 0., 0., 0., 0.]
                #Tgi = 1
                for Tgi in range(1,5):
                    tp = 0
                    ti = 1
                    #if T_gate[Tgi] < Duration[scn]:
                    for t in TVD[1:,0]:                
                        if  tp < T_gate[Tgi] and T_gate[Tgi] <= t:
                            #t = TVD[ti,0]
                            M_gate[Tgi] = TVD[ti,1]
                            RR_gate[Tgi] = TVD[ti,2]
                            #print("ti={:5d} Mass {:12.2f} ({:12.2f}) \& Release Rate {:12.2f} ({:12.2f})at {:12.2f} read".format(ti,M_gate[Tgi],TVD[ti,1],RR_gate[Tgi],TVD[ti,2],t))
                        tp = t
                        ti = ti+1
                #print("{:12d}{:12d}{:12d}{:12d}{:12d}".format(T_gate[0],T_gate[1],T_gate[2],T_gate[3],T_gate[4]))
                #print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(M_gate[0],M_gate[1],M_gate[2],M_gate[3],M_gate[4]))
                #print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(RR_gate[0],RR_gate[1],RR_gate[2],RR_gate[3],RR_gate[4]))

                #Find an Event that fits the 'path' + 'weather'
                # i=0
                # while (not (lEvent[i].Path == path and lEvent[i].Weather == weather)):
                #     i = i + 1
                #     if i == len(lEvent):
                #         break
                # if i < len(lEvent):
                #     key = path + "\\" + weather
                #     #Save the results of T, M & RR
                #     lEvent[i].Discharge.Ts = T_gate
                #     lEvent[i].Discharge.Ms = M_gate
                #     lEvent[i].Discharge.RRs = RR_gate
                #     lEvent[i].TVDRead = True
                #     lEvent[i].TVD = TVD
                #     print("TVD Read & Saved : {} {}".format(path, weather))
                # else:
                #     print("Event corresponding to {} {} not found".format(path, weather))

                for e in lEvent:
                    if e.Path == path:
                        e.Discharge.Ts = T_gate
                        e.Discharge.Ms = M_gate
                        e.Discharge.RRs = RR_gate
                        e.TVDRead = True
                        e.TVD = TVD
                        print(e.Path,e.Weather, "TVD Read")
                



                if TVDPLot == True:
                    fig,ax1 = plt.subplots()                
                    Time = TVD[:,0]
                    Mass = TVD[:,1]
                    RelRate = TVD[:,2]

                    masscolor = 'tab:blue'
                    ax1.set_xlabel('Time [s]')
                    ax1.set_ylabel('Mass [kg]',color=masscolor)
                    ax1.plot(Time,Mass,color=masscolor)
                    ax1.set_ylim(bottom=0)
                    ax1.tick_params(axis='y',labelcolor=masscolor)
                    # ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
                    # ax1.set_xlim(left=0, right=3600)
                    
                    ax2 = ax1.twinx()
                    rrcolor = 'tab:red'
                    ax2.set_ylabel('Release Rate [kg/s]',color=rrcolor)
                    ax2.plot(Time,RelRate,color=rrcolor)
                    ax2.set_ylim(bottom=0)
                    ax2.tick_params(axis='y',labelcolor=rrcolor)
                    pngfilename = "TVD-"+pv+"_"+hole
                    plt.title(pngfilename)            
                    plt.show()
                    fn = slugify(pngfilename)
                    fig.savefig(".\\tvd_rev.B\\{}.png".format(fn))
                    plt.close()
                    

                #Copy the discharge pattern from 6.6/D to 1.2/F
                # i=0
                # while (not (lEvent[i].Path == path and lEvent[i].Weather == '1.2/F')):
                #     i = i + 1
                #     if i == len(lEvent):
                #         break
                # if i < len(lEvent):
                #     key = path + "\\" + '1.2/F'
                #     lEvent[i].Discharge.Ts = T_gate
                #     lEvent[i].Discharge.Ms = M_gate
                #     lEvent[i].Discharge.RRs = RR_gate
                #     lEvent[i].TVDRead = True
                #     lEvent[i].TVD = TVD
                #     #print(lEvent[i].Discharge.Ms)
                # else:
                #     print("Event corresponding to {} {} not found".format(path, '1.2/F'))                
        else:
            print("skip line {}".format(r))
            # continue
        r = r + 1

#End of reading Discharge files

#For an Event, find the relevant Leak and return its frequency
#r = 10
#key = lEvent[r].Path
#print("Path: {}  Frequency: {} Release Height: {}".format(key,Frequency[key],ReleaseHeight[key]))

#Reading Jet fire
#1. Find an Event that has the same study_folder_pv, scenario, & weather
# lEvent[##].JetFire = JetFiire
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shJet.max_row+1):
    study_folder_pv = shJet.cell(r,1).value
    scenario = shJet.cell(r,2).value
    weather = shJet.cell(r,3).value
    path = study_folder_pv + "\\" + scenario
        
    study,pv,folder = study_folder_pv.split("\\")
    # holesizeindex = ['SM','ME','MA','LA'].index(scenario[0:2])

    key = pv + "\\" + scenario + "\\" + weather
    #Find the relevant event
    # for e in lEvent:
    #     if e.Key == key: 
    #         break
    
    i=0
    while not (lEvent[i].Path == path and lEvent[i].Weather == weather):
        i = i + 1
        if i == len(lEvent):
            break
    if i < len(lEvent):
        # key = path + "\\" + weather
        if weather in weathers:
            jfl = shJet.cell(r,10).value
            #J00[key] = jfl
            sep = shJet.cell(r,11).value
            JetSEP[key] = sep

            e = lEvent[i]
            jff = e.Frequency*e.PImdIgn/NumDirections
            # if weather == '1.2/F':
            #     jff = jff*0.18
            # elif weather == '6.6/D':
            #     jff = jff*0.82
            JetFrequency[key] = jff
            lEvent[i].JetFire = JetFire(jfl,sep,jff)

            if e.Discharge.Ts[2] != 0:
                jf0 = e.JetFire.Length
                m0 = e.Discharge.RRs[0]
                jf0fit = jffit(m0)
                jfscale = e.JetFire.Length/jf0fit
                e.jfscale = jfscale
                lEvent[i].JetFire.JetLengths[0] = jf0
                for ti in range(1,5):
                    #t = a_event.Ts[i]
                    m = e.Discharge.RRs[ti]
                    jf = jffit(m)
                    lEvent[i].JetFire.JetLengths[ti] = jf*jfscale
                    #lEvent[i].JetFire.JetLengths[ti] = jf
                J00[key] = jf0
                J05[key] = lEvent[i].JetFire.JetLengths[1]
                J10[key] = lEvent[i].JetFire.JetLengths[2]
                J30[key] = lEvent[i].JetFire.JetLengths[3]
                J60[key] = lEvent[i].JetFire.JetLengths[4]

        else:
            print("{} Jet fire events read".format(k))
            break
    # k=k+1
#Read Jet, shJet


#shExplosion
ExpD1 = {}
ExpD2 = {}
ExpD3 = {}
ExpFreq = {}
#Read Early Pool Fire
for r in range(2,shExplosion.max_row+1):
    study_folder_pv = shExplosion.cell(r,1).value
    scenario = shExplosion.cell(r,2).value
    weather = shExplosion.cell(r,3).value
    # path = study_folder_pv + "\\" + scenario    
    # key = path + "\\" + weather    
    p1 = shExplosion.cell(r,9).value    
    p2 = shExplosion.cell(r,10).value    
    p3 = shExplosion.cell(r,11).value    
    d1 = shExplosion.cell(r,12).value    
    d2 = shExplosion.cell(r,13).value    
    d3 = shExplosion.cell(r,14).value    

    study,pv,folder = study_folder_pv.split("\\")
    key = pv + "\\" + scenario + "\\" + weather

    ExpD1[key] = d1
    ExpD2[key] = d2
    ExpD3[key] = d3
    
    for e in lEvent:
        if e.Key == key:
            # fexp = e.Frequency*(1-e.PImdIgn)*e.PDelIgn*e.PExp_Ign
            fexp = e.Frequency*e.PDelIgn*e.PExp_Ign
            ExpFreq[key] = fexp
            e.Explosion = Explosion(fexp,[p1,p2,p3],[d1,d2,d3])


#shDispersion
DistLFLFrac = {}
DistLFL = {}
DispFreq = {}
#Read Flammable Dispersion
for r in range(2,shDispersion.max_row+1):
    study_folder_pv = shDispersion.cell(r,1).value
    scenario = shDispersion.cell(r,2).value
    weather = shDispersion.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    # key = path + "\\" + weather    
    # key = pv+ "\\" + hole + "\\" + weather    

    study,pv,folder = study_folder_pv.split("\\")
    key = pv + "\\" + scenario + "\\" + weather
    # 15,16,17
    dfmax = shDispersion.cell(r,15).value    
    dfmin = shDispersion.cell(r,16).value    
    wf = shDispersion.cell(r,17).value    
    # 19,20,21
    dmax = shDispersion.cell(r,19).value    
    dmin = shDispersion.cell(r,20).value    
    w = shDispersion.cell(r,21).value          

    ffmax = shDispersion.cell(r,27).value          
    ffw = shDispersion.cell(r,29).value          
    
    
    DistLFLFrac[key] = dfmax
    DistLFL[key] = dmax
    
    for e in lEvent:
        if e.Key == key:
            # fdisp = e.Frequency*(1-e.PImdIgn)*e.PDelIgn*(1-e.PExp_Ign)
            fdisp = e.Frequency*(1-e.PImdIgn - e.PDelIgn)
            DispFreq[key] = fdisp
            e.Dispersion = Dispersion(fdisp,dfmax,dfmin,wf,dmax,dmin,w,ffmax,ffw)

if 'Early Pool Fire' in cExl.sheetnames:
    shPool = cExl['Early Pool Fire']
    EPSEP = {}
    EPFreq = {}
    #Read Early Pool Fire
    for r in range(2,shPool.max_row+1):
        study_folder_pv = shPool.cell(r,1).value
        scenario = shPool.cell(r,2).value
        weather = shPool.cell(r,3).value
        path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather

        sep = shPool.cell(r,11).value    
        EPSEP[key] = sep
        dia = shPool.cell(r,10).value    
        for e in lEvent:
            if e.Key == key:
                epff = e.Frequency*e.PImdIgn
                EPFreq[key] = epff
                e.EarlyPoolFire = EarlyPoolFire(dia,sep,epff)
                break



if 'Late Pool Fire' in cExl.sheetnames:
    shPool = cExl['Late Pool Fire']
    LPSEP = {}
    LPFreq = {}
    #Read Late Pool Fire
    for r in range(2,shPool.max_row+1):
        study_folder_pv = shPool.cell(r,1).value
        scenario = shPool.cell(r,2).value
        weather = shPool.cell(r,3).value
        # path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather
        sep = shPool.cell(r,11).value    
        LPSEP[key] = sep
        dia = shPool.cell(r,10).value    
        for e in lEvent:
            if e.Key == key:
                # epff = e.Frequency*(1-e.PImdIgn)*e.PDelIgn
                epff = e.Frequency*e.PDelIgn
                LPFreq[key] = epff
                e.LatePoolFire = LatePoolFire(dia,sep,epff)


#shFireball
if 'Fireball' in cExl.sheetnames:
    shFireball = cExl['Fireball']
    DiaFireball = {}
    SEPFireball = {}
    FireballFreq = {}
    #Read Flammable Dispersion
    for r in range(2,shFireball.max_row+1):
        study_folder_pv = shFireball.cell(r,1).value
        scenario = shFireball.cell(r,2).value
        weather = shFireball.cell(r,3).value
        # path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather
        
        dia = shFireball.cell(r,9).value    
        sep = shFireball.cell(r,10).value    
        
        DiaFireball[key] = dia
        SEPFireball[key] = sep
        
        for e in lEvent:
            if e.Key == key:
                ffireball = e.JetFire.Frequency
                FireballFreq[key] = ffireball
                e.Fireball = Fireball(ffireball,dia,sep)


with open(element_dump_filename,'wb') as element_dump:
    dill.dump(lEvent,element_dump)



F = Fe = Fl = 0.
for e in lEvent:
    if "-L" in e.Key:
        F += e.Frequency
    if e.EarlyPoolFire is not None:    
        Fe += e.EarlyPoolFire.Frequency
    if e.LatePoolFire is not None:    
        Fl += e.LatePoolFire.Frequency

print("{:15s} Liquid Leak Frequency: {:.2e}".format(Area, F))
print("{:15s} Pool fire Toal Frequency: Early {:.2e} Late {:.2e}".format(Area, Fe,Fl))
# print("{:15s} Pool fire Toal Frequency: {:.2e}".format(Area, sum(EPFreq.values()) + sum(LPFreq.values())))
