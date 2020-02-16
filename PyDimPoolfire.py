
from openpyxl import load_workbook
import math
import numpy as np
#import dill

iExlFilename='h11c4_dim_pool_fire'
Modules = ['CDA','KOD','MeOHS','HFP','HFS','HAP','HAS']
element_dump_filename = 'h11_dump_jet'

iExlFilename='v04c_dim_pool_fire'
Modules = ['P02','P03','P04','P05','S02','S03','S04','S05']
element_dump_filename = 'v09_dump_jet'



iExl=load_workbook(filename=iExlFilename+'.xlsx')
shtPool = iExl['Sheet1']


#with open(element_dump_filename,'rb') as elements_dump:
    #lEvent = dill.load(elements_dump)  

#for r in range(6,numscn):
numrows = shtPool.max_row
Dxx = []
for r in range(2,numrows+1):
    module = shtPool.cell(r,1).value
    path = shtPool.cell(r,2).value
    #area, pv, eq_hole_weather = path.split("\\")
    area = "Topsisde"
    pv, eq_hole_weather = path.split("\\")
    weather = eq_hole_weather[-5:]
    hole = eq_hole_weather[-8:-6]
    eq = eq_hole_weather[:-7]
    
    rr = shtPool.cell(r,3).value
    t2 = shtPool.cell(r,4).value
    mass_split = shtPool.cell(r,5).value
    burn_rate = shtPool.cell(r,6).value
    D0 = shtPool.cell(r,7).value
    H0 = shtPool.cell(r,8).value
    Te = shtPool.cell(r,9).value
    D00 = shtPool.cell(r,10).value
    D05 = shtPool.cell(r,11).value
    D10 = shtPool.cell(r,12).value
    D30 = shtPool.cell(r,13).value
    D60 = shtPool.cell(r,14).value
    Freq = shtPool.cell(r,15).value
    Dxx.append([module,Freq, D00, D05, D10, D30, D60])

#Area, PV, Equip, X, Y, Height, Module, Deck Level, Hole, Wether, PoolFireFrequency, SEP, D00, D05,D10,D30,D60



RRThreshold = 0.1
Jsets = {'00':J00, '05':J05,'10':J10,'30':J30,'60':J60}
RRIndices = {'00':0,'05':1,'10':2,'30':3,'60':4}
#load JetAnalysis.py
for Threshold in ['00', '05', '10','30','60']:
# for Threshold in ['05']:
    RRIndex = RRIndices[Threshold]
       
    #For each module
    for mod in Modules:    
    # for mod in ['P02']:
        # Js = Dxx[:,[0,RRIndices[Threshold]+2]]
        nr = 0
        Js = np.array([0,0])
        for r in range(0,numrows-1):
            if Dxx[r][0] == mod:
                Js = np.vstack([Js, [Dxx[r][1], Dxx[r][RRIndex+2]]])
                nr += 1
        if nr > 0:
            cJs =sorted(Js,key= lambda f1:f1[1])
            # Js = sorted(jd, key = lambda fl: fl[2])
            DimFound = False
            for ir in range(nr-2,-1,-1):
                cp = cJs[ir+1][0]
                cJs[ir][0] += cp
                cn = cJs[ir][0]       
                if cp > 1E-4 and DimFound == False:
                    jp = cJs[ir+1][1]
                    j0 = jp                    
                    print('{}: {}: Dimensioning Pool Diameter {:8.1f}'.format(Threshold,mod,j0))
                    DimFound = True
                    # break
                if cn >= 1E-4 and cp < 1E-4 and DimFound == False:
                    jp = cJs[ir+1][1]
                    jn = cJs[ir][1]
                    j0 = (jn-jp)/(cn-cp)*(1E-4-cp) + jp
                    print('{}: {}: Dimensioning Pool Diameter {:8.1f}'.format(Threshold,mod,j0))
                    DimFound = True
                    # break
                """ else:
                    print('{}: {}: No release larger than {}'.format(Threshold, mod, RRThreshold))             """
            if ir == 0 and DimFound == False:
                print('{}: {}: No dimensioning jet length'.format(mod, Threshold))
        else:
            print("No data for "+mod)
        """  CF = cJs[:,0]JFL = ec[:,1]

        fig,ax1 = plt.subplots()
        masscolor = 'tab:blue'
        ax1.set_xlabel('Pool Diameter [m]')
        ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
        ax1.semilogy(JFL,CF,color=masscolor)
        ax1.set_ylim(bottom=1E-6)
        ax1.tick_params(axis='y',labelcolor=masscolor)
        ax1.grid(True,which="major")
              
        
        plt.title("Pool Diameter: " +mod+" at "+Threshold+" min")            
        plt.show()
        fig.savefig("{}T{}.png".format(mod,Threshold))
        plt.close() """


        

