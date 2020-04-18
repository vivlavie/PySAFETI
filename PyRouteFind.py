

#PyRouteFind.py 


#To read heat on and around SCE's
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import numpy as np
# import dill
import unicodedata
import re
import csv
import os
import time
from kfxtools import * #fit for Python 2.

Ns = {}
with open('nodecords.csv', 'r') as file:
# with open('n.csv', 'r') as file:
    reader = csv.reader(file)
    for n in reader:
        Ns[n[0]] = [float(x) for x in n[1:]]
        # print(Ns[n[0]])
Cs = []
with open('connectivity.csv', 'r') as file:
# with open('c.csv', 'r') as file:
    reader = csv.reader(file)
    for c in reader:
        Cs.append([c[0],c[1],True])


Vh = 1.3 #average
Vv = 0.49 #downward slowest person

R = {}    
for c in Cs:
    x1,y1,z1 = Ns[c[0]]
    x2,y2,z2 = Ns[c[1]]
    if (z1 == z2):
        v = Vv
    elif (x1 == x2) and (y1 == y2) and: not (z1 == z2):
        v = Vv
       
    d = math.sqrt((x2-x1)^2+(y2-y1)^2+(z2-z1)^2)
    s = d / v
    if not (c[0] in R.keys()):
        R[c[0]] = {c[1]:s}
    else:
        R[c[0]][c[1]] = s

    if not (c[1] in R.keys()):
        R[c[1]] = {c[0]:s} 
    else:
        R[c[1]][c[0]] = s

import heapq

# 탐색할 그래프와 시작 정점을 인수로 전달받습니다.
def dijkstra(graph, start, end):
    # 시작 정점에서 각 정점까지의 거리를 저장할 딕셔너리를 생성하고, 무한대(inf)로 초기화합니다.
    distances = {vertex: [float('inf'), start] for vertex in graph}
    # 그래프의 시작 정점의 거리는 0으로 초기화 해줌
    distances[start] = [0, start]
    # 모든 정점이 저장될 큐를 생성합니다.
    queue = []
    # 그래프의 시작 정점과 시작 정점의 거리(0)을 최소힙에 넣어줌
    heapq.heappush(queue, [distances[start][0], start])
    while queue:        
        # 큐에서 정점을 하나씩 꺼내 인접한 정점들의 가중치를 모두 확인하여 업데이트합니다.
        current_distance, current_vertex = heapq.heappop(queue)
        # 더 짧은 경로가 있다면 무시한다.
        if distances[current_vertex][0] < current_distance:
            continue            
        for adjacent, weight in graph[current_vertex].items():
            distance = current_distance + weight
            # 만약 시작 정점에서 인접 정점으로 바로 가는 것보다 현재 정점을 통해 가는 것이 더 가까울 경우에는
            if distance < distances[adjacent][0]:
                # 거리를 업데이트합니다.
                distances[adjacent] = [distance, current_vertex]
                heapq.heappush(queue, [distance, adjacent])    
    path = end
    path_output = end + '->'
    while distances[path][1] != start:
        path_output += distances[path][1] + '->'
        path = distances[path][1]
    path_output += start
    print (path_output)
    return distances 


r_th = [37500, 12500, 4700]
t_th = [1, 40, 120]
impairment = ['Imd fatality','12.5kW/mw','4.7kW/m2']
fieldnum = 0

# P04_A_AP        : /projects/300341/Reliance/Fire/P04/J01
# P04_B_FS        : /projects/300341/Reliance/Fire/P04/J06
# S05_A_F           : /projects/300341/Reliance/Fire/S05/J09
# S05_B_F           : /projects/300341/Reliance/Fire/S05/J12
# S04_A_A          : /projects/300341/Reliance/Fire/S04/J13
# S03_A_P          : /projects/300341/Reliance/Fire/S03/J18
# S03_B_A          : /projects/300341/Reliance/Fire/P04/J19
NumRowsJetInfo = 5
basefolder = "./Rev.B/"
Js = {}
Js['J01'] = 'P04_A_AP'
Js['J02'] = 'P04_A_FP'
Js['J03'] = 'P04_A_AS'
Js['J04'] = 'P04_A_FS'
Js['J05'] = 'P04_B_AS'
Js['J06'] = 'P04_B_FS'
Js['J07'] = 'S05_A_AS'
Js['J08'] = 'S05_A_AP'
Js['J09'] = 'S05_A_F'
Js['J10'] = 'S05_A_F_10'
Js['J11'] = 'S05_B_AP'
Js['J12'] = 'S05_B_F'
Js['J13'] = 'S04_A_A'
Js['J14'] = 'S04_A_FP'
Js['J15'] = 'S04_A_FS'
Js['J16'] = 'S04_B_AP'
Js['J17'] = 'S04_B_F'
Js['J18'] = 'S03_A_P'
Js['J19'] = 'S03_B_A'
Js['J20'] = 'S03_B_F'
Js['J21'] = 'P02_B_FS'
Js['J22'] = 'P02_B_FP'
Js['J23'] = 'P05_A_P'
Js['J24'] = 'P05_A_S'
Js['J25'] = 'P05_B_A'
Js['J26'] = 'P05_B_F'
Js['J27'] = 'P03_B_P'
Js['J28'] = 'P03_B_FS'
Js['J29'] = 'KOD_B'


# for j in Js.keys():
for j in ['J02','J26']:
    colid = int(j[-2:])+10
    fdr = Js[j][:3]
    # fn = basefolder + fdr + "/" + j + "/" + j+"_rad_exit.r3d"    
    # fnv = basefolder + fdr + "/" + j + "/" + j+"_vis_exit.r3d"    
    fn = basefolder + "/" + j+"_rad_exit.r3d"    
    fnv = basefolder +"/" + j+"_vis_exit.r3d"    
    
    if (os.path.exists(fn) == False):
        print(fn + " does not exist")
    elif (os.path.exists(fnv) == False):        
        print(fnv +" does not exist")
    else:    
        print(fn)
        T = readr3d(fn)    
        Tv = readr3d(fnv)    
        #Define field name to read in
        fieldname = T.names[fieldnum]
        fieldnamev = Tv.names[fieldnum]
        print(fieldname)
        print(Js[j],fn,fieldname)
        er_imp_kfx = open(j+"_er_imp.kfx","w")

        #For each connection
        for c in Cs:
            x1,y1,z1 = Ns[c[0]]
            x2,y2,z2 = Ns[c[1]]
            z1 += 1.5        
            z2 += 1.5        
            dx,dy,dz = x2-x1,y2-y1,z2-z1
            x1m,y1m,z1m = x1+0.25*dx,y1+0.25*dy,z1+0.25*dz
            xm,ym,zm = x1+0.5*dx,y1+0.5*dy,z1+0.5*dz
            xm2,ym2,zm2 = x1+0.75*dx,y1+0.75*dy,z1+0.75*dz
            
            
            
            #Impairment assessment w.r.t. Radiation
            r1 = T.point_value(x1,y1,z1,fieldnum)
            r1m = T.point_value(x1m,y1m,z1m,fieldnum)
            rm = T.point_value(xm,ym,zm,fieldnum)
            rm2 = T.point_value(xm2,ym2,zm2,fieldnum)
            r2 = T.point_value(x2,y2,z2,fieldnum)
            #Personnel moving speed, horizontal or vertical
            if abs(dz) < 0.01:
                V = Vh
            else:
                V = Vv
            dt = sqrt(dx*dx+dy*dy+dz*dz)/V
            ravg = (r1+r1m+rm+rm2+r2)/5.
            color = "0 1 0"
            #For each radiation impairment criteria
            for ck in range(0,len(r_th)):
                if (ravg > r_th[ck]) and (dt > t_th[ck]):
                    print ( Ns[c[0]], "->",Ns[c[1]], "{:8.1f} kW/m2 {:6.1f} sec {:s}".format(ravg/1000, dt, impairment[ck]))
                    c[2] = False #Set that the connnection is 'impaired'
                    color = "1 0 0"
                    #Python 2.x
                    # er_imp_kfx.write("COLOR: 1 0 0\n PART: "+c[0]+" "+c[1]+"\n")
                    #Python 3.x
                    # print("COLOR: 1 0 0", file = er_imp_kfx)                                   
            
            
            #Impairment assessment w.r.t. Visibility
            r1v = Tv.point_value(x1,y1,z1,fieldnum)
            r1mv = Tv.point_value(x1m,y1m,z1m,fieldnum)
            rmv = Tv.point_value(xm,ym,zm,fieldnum)
            rm2v = Tv.point_value(xm2,ym2,zm2,fieldnum)
            r2v = Tv.point_value(x2,y2,z2,fieldnum)
            # print(r1v, r1mv, rmv, rm2v, r2v)
            # Visibility = (r1v+r1mv+rmv+rm2v+r2v)/5.
            Visibility = max([min([r1v,r1mv,rmv,rm2v,r2v]),1])
            if (ravg > 4700) and (Visibility < 5.):
                    print ( Ns[c[0]], "->",Ns[c[1]], "Min visibility {:8.1f}".format(Visibility))                    
                    print([r1v,r1,r1mv,r1m,rmv,rm,rm2v,rm2,r2v,r2])
                    if c[2] == False:
                        #impaired by both heat and visibility
                        color = "0 1 1"           
                    else:
                        #impaired only by visibility
                        color = "0 0 1"           
                    c[2] = False #Set that the connnection is 'impaired'


            er_imp_kfx.write("COLOR: "+color+"\nPART: "+c[0]+"_"+c[1]+"\n")                                
            er_imp_kfx.write("SBOX: {:10.1f} {:10.1f} {:10.1f} {:10.1f} {:10.1f} {:10.1f} 400 400\n".format(x1*1000,y1*1000,z1*1000,x2*1000,y2*1000,z2*1000))
        er_imp_kfx.close()

""" 
        for k in SCE.keys():        
        # for k in ['020-VZ-002']:        
            # if "VZ" in k:
                s = SCE[k]
                x = s[0]
                y = s[1]
                z = s[2]
                Ori = s[3]
                D = s[4]
                H = s[5]
                sce_name = s[8]
                rowid = s[10]
                if Ori == 'X':
                    dx = H/2.
                    dy = D/2.
                    dz = dy
                elif Ori == 'Y':
                    dy = H/2.
                    dx = D/2.
                    dz = dx
                elif Ori == 'Z':
                    dz = H/2.
                    dy = D/2.
                    dx = dy
                elif Ori == 'C':
                    dx = dy = dz = D/2.
                if Ori != None:
                    Xs = [x - 1.2*dx, x, x + 1.2*dx]
                    Ys = [y - 1.2*dy, y, y + 1.2*dy]
                    Zs = [z - 0.4*dz, z, z + 2.0*dz]
                    r = 0.
                    rmax = 0.
                    rsum = 0.
                    ncount = 0
                    ravg = 0.
                    for xi in Xs:
                        for yi in Ys:
                            if Ori != "C":
                                for zi in Zs:
                                    r = T.point_value(xi,yi,zi,fieldnum)                                
                                    # print("{:8.1f}{:8.1f}{:8.1f}{:8.1e}{:8.1f}".format(xi,yi,zi,r/1000,rmax/1000))
                                    if r > 1000:
                                        rmax = max(r,rmax)
                                        ncount += 1
                                        rsum += r
                                # ravg = rsum/(len(Xs)*len(Ys)*len(Zs)-1) #neglecting the reading at the center, that shall be 0!
                            else: #for the crane only!                    
                                for zi in [z]:
                                    r = T.point_value(xi,yi,zi,fieldnum)
                                    # print("{:8.1f}{:8.1f}{:8.1f}{:8.1e}{:8.1f}".format(xi,yi,zi,r/1000,rmax/1000))
                                    if r > 1000:
                                        rmax = max(r,rmax)
                                        ncount += 1                                    
                                        rsum += r
                    if (ncount > 0) and (rsum > 0.01):
                        ravg = rsum/ncount
                        # print("#{:20s} {:6s} {:4s} {:6.1f} {:6.1f} {:10d}".format(k,s[6],s[7],rmax/1000,ravg/1000,ncount))
                        # print("BOX: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} #{:15s} {:50s} {:6.1f} {:6.1f} {:2d}".format((Xs[0])*1000,(Ys[0])*1000,(Zs[0])*1000,(Xs[2]-Xs[0])*1000,(Ys[2]-Ys[0])*1000,(Zs[2]-Zs[0])*1000,k,sce_name,rmax/1000,ravg/1000,ncount))
                        # shHeatMax.cell(rowid,colid).value = "{:8.2f}".format(rmax/1000)
                        # shHeatAvg.cell(rowid,colid).value = "{:8.2f}".format(ravg/1000)
                        shHeatMax.cell(rowid+NumRowsJetInfo,colid).value = rmax/1000
                        shHeatAvg.cell(rowid+NumRowsJetInfo,colid).value = ravg/1000
                        
                    # else:
                        # print(k,"something wrong if the following two numbers are not zero", ncount, rsum)        """                         

# iHeat.save('SCE_Heat_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')