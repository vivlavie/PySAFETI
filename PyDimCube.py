#PyDimCube
#For a cube, read from each cube the duration that jet fire impinges on it

from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill

import matplotlib.pylab as pltparam

def DtoH (D05):
    if D05 > 19.7:
            H05 = 0.4664*D05 + 18.345
    elif D05 < 1.0:
        H05 = 0.
    else: #  1.0 < D05 < 19.7:
        H05 = 1.2592*D05 + 2.7235
    return H05
def PD(t,De,Te):
    if t < Te:
        PD = De * (1-math.sqrt(t/Te))
    else:
        PD = 0
    return PD
#pool burning rate 0.062 kg/m2-s
#MS ; spilt mass
#Pool fire duration
# Ms / (3.14*D^2/4 * br) *3/2 where 3/2 is a factor to consider shringking pool


pltparam.rcParams["figure.figsize"] = (8,3)
pltparam.rcParams['lines.linewidth'] = 2
# pltparam.rcParams['lines.color'] = 'r'
pltparam.rcParams['axes.grid'] = True 

# element_dump_filename = 'v08_element.bin'
""" with open(element_dump_filename,'wb') as element_dump:
    dill.dump(lEvent,element_dump) """
with open('v08_element.bin','rb') as v08dump:
    lEvent = dill.load(v08dump)

with open('v09_dump_jet','rb') as v09dump:
    lEvent = dill.load(v09dump)

iExlFilename='v08_9mi'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    folder  = shPV.cell(r,3).value
    pv  = shPV.cell(r,15).value
    key = study  + "\\" +  folder  + "\\" + pv
    X[key] = shPV.cell(r,152).value
    Y[key] = shPV.cell(r,153).value
    r += 1
#Read X&Y and Put into lEvent
for e in lEvent:
    e.X = X[e.Study_Folder_Pv]
    e.Y = Y[e.Study_Folder_Pv]




#construct list of PV's
pvloc = []
listPV = X.keys()
for pv in listPV:
    if pvloc == []:
        pvloc = [X[pv],Y[pv]]
    else:
        pvloc = np.vstack([pvloc,[X[pv],Y[pv]]])
##nd of preparations for controus drawing


Resolution = 2
KK = np.arange(0,5) #Index for each time threshold to be read for each jet length
XX = np.linspace(99,224,50)
YY = np.linspace(-27,27,21)
# TimeIndices = ['00','05','10','30','60']
# KK = [1, 2, 3, 4]

icubeloc='SCE_CUBE_XYZ'
iExl=load_workbook(filename=icubeloc+'.xlsx')
shCube = iExl['xyz']
ncube = shCube.cell(1,1).value
Xcube = {}
Ycube = {}
Zcube = {}
Cubes = []
CubeImpingeDuration = {}
for i in range(3,ncube+3):
    id = shCube.cell(i,1).value
    Cubes.append(id)
    Xcube[id] = shCube.cell(i,2).value
    Ycube[id] = shCube.cell(i,3).value
    Zcube[id] = shCube.cell(i,4).value

for i in range(0,ncube):    
    li = 0
    sep = np.zeros((len(lEvent),2))
    ImpingeDurationArray = []
    id = Cubes[i]
    xx = Xcube[id]
    yy = Ycube[id]
    zz = Zcube[id]
    for e in lEvent:
        #Exposure to jet fire
        #read frequency
        sep[li,1] = e.JetFire.Frequency
        #check if the receptor is in the flame
        dx = xx - e.X
        dy = yy - e.Y
        dz = zz - (e.ReleaseHeight+35)
        rr = math.sqrt(dx*dx+dy*dy+dz*dz)
        #Array of jet length
        jl_e = 2.8893*np.power(55.5*e.TVD[:,2],0.3728) #the 3rd column of the matrix, Lowesmith formulat
        t_e = interp1d(jl_e,e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'
        if max(jl_e) < rr:                
            sep[li,0] = 0 #JF doesn't reach the cube
        elif min(jl_e) > rr:
            sep[li,0] = e.TVD[-1,0] #JF keeps impinging the cube through the release
        else:
            sep[li,0] = t_e(rr)
        # sep[li,1] = e.Frequency
        ImpingeDurationArray.append([sep[li,0],sep[li,1],e.Key+"_Jet"])

        #Exposure to pool fire
        if e.EarlyPoolFire != None:
            #Read pool fire duration, ho
            dd = e.EarlyPoolFire.Diameter
            Ms = e.Discharge.Ms[0] - e.Discharge.Ms[-1]
            PFD = Ms/(3.14*dd*dd/4 * 0.062)*3/2 #Pool Fire Duration
            rr = math.sqrt(dx*dx+dy*dy)
            if (rr < dd) & (zz > e.ReleaseHeight):
                di = PFD*(dd-rr)/dd
                # di = min ( (rr-dd)/V_pool_shrink, (zz-e.ReleaseHeight)/(V_pool_shrink*) )
            ImpingeDurationArray.append([di,e.EarlyPoolFire.Frequency,e.Key+"_EarlyPool"])
        if e.LatePoolFire != None:
            #Read pool fire duration, ho
            dd = e.LatePoolFire.Diameter
            Ms = e.Discharge.Ms[0] - e.Discharge.Ms[-1]
            PFD = Ms/(3.14*dd*dd/4 * 0.062)*3/2 #Pool Fire Duration
            rr = math.sqrt(dx*dx+dy*dy)
            if (rr < dd) & (zz > e.ReleaseHeight):
                di = PFD*(dd-rr)/dd
                # di = min ( (rr-dd)/V_pool_shrink, (zz-e.ReleaseHeight)/(V_pool_shrink*) )
            ImpingeDurationArray.append([di,e.LatePoolFire.Frequency,e.Key+"_LatePool"])

        li += 1            
    #To pin-point a scenario that give the dimensioning scenario
    IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0]) #with the longest duration at the bottom
    
    cf = 0
    jp = IDAsorted[-1]
    DimFreq = 1.0E-4
    di = 0
    cfp = jp[1]#frequency for the longest duration jet fire
    InterpolationSuccess = False
    if cfp > DimFreq:
        di = jp[0]
        scn = jp[2]
    else:
        for j in IDAsorted[-2:0:-1]:
            cf = cfp + j[1]
            if cf >= DimFreq and cfp < DimFreq:
                dp = jp[0]
                dn = j[0]
                di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
                # print('Dimensioning jet duration {:8.1f}'.format(di))
                scnp = jp[2]
                scn = j[2]
                InterpolationSuccess = True
                break
            cfp = cf
            # print(cfp,cf)
            jp = j
        
            # break        
    """ # Sort 'sep' on the SEP
    # Read 'setp' corresponding to 1E-4
    # sep.sort(column=1)
    # sep_sorted = np.sort(sep,axis=1)[::-1] #sort by SEP in the descending order
    sep_sorted = sep[sep[:,0].argsort()][::-1]
    f_max2min = sep_sorted[:,1] #extract frequency
    cf = np.cumsum(f_max2min)
    # sep_sorted_with_cf = sep_sorted
    # sep_sorted_with_cf[:,1] = cf
    #interpolate on 'Column 2' and read from 'Column 1'
    if cf[-1] > 1E-4:
        if cf[0] < 1E-4:
            fl = interp1d(cf,sep_sorted[:,0],kind='linear')
            sep_dim = fl(1E-4)
        else:
            # print("Impinged longer than 1 hr ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[0]))    
            sep_dim = sep_sorted[0,0] #should be 3,600
    else:
        print("Not dimensioning at ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[-1]))
        sep_dim = 0.            
    # ImpingeDuration[i,j] = sep_dim
    CubeImpingeDuration[Cubes[i]] = sep_dim
    print("Cube: {} {:4.1f} {:4.1f} - Scenario: {}".format(id,di,sep_dim,scn))
     """
    CubeImpingeDuration[Cubes[i]] = di
    ts,pv,scn,hole,weather = scn.split("\\")
    dmy,weather_fire = weather.split(" ")
    hole,tlv = hole.split("_")
    print("Cube: {:20s} {:6s} fire from {:20} during {:8.1f} [sec]".format(id,weather_fire[6:],di,scn+"_"+hole+"_"+weather_fire[:5]))


for c in Cubes:
    print("{:20s} {:8.1f} [min]".format(c,CubeImpingeDuration[c]/60))

Levels = [1.5,10.5]
ImpingeDuration = np.zeros([len(XX),len(YY),len(Levels)])
#For x & y contours of duration
for k in range(0,len(Levels)):
    zz = Levels[k]
    for i in range(0,len(XX)):
        for j in range(0,len(YY)):
            li = 0
            sep = np.zeros((len(lEvent),2))
            ImpingeDurationArray = []
            for e in lEvent:
                #read frequency
                if e.ReleaseHeight <= zz:
                    sep[li,1] = e.JetFire.Frequency
                    #check if the receptor is in the flame
                    dx = XX[i]-e.X
                    dy = YY[j]-e.Y
                    dz = (zz-e.ReleaseHeight)
                    rr = math.sqrt(dx*dx+dy*dy+dz*dz)
                    #Array of jet length
                    jl_e = 2.8893*np.power(55.5*e.TVD[:,2],0.3728) #the 3rd column of the matrix, Lowesmith formulat
                    t_e = interp1d(jl_e,e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'
                    if max(jl_e) < rr:                
                        sep[li,0] = 0
                    elif min(jl_e) > rr:
                        sep[li,0] = e.TVD[-1,0]
                    else:
                        sep[li,0] = t_e(rr)
                    sep[li,1] = e.Frequency
                    ImpingeDurationArray.append([sep[li,0],sep[li,1],e.Key])
                    li += 1            
            #To pin-point a scenario that give the dimensioning scenario
            IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0])

            # Sort 'sep' on the SEP
            # Read 'setp' corresponding to 1E-4
            # sep.sort(column=1)
            # sep_sorted = np.sort(sep,axis=1)[::-1] #sort by SEP in the descending order
            sep_sorted = sep[sep[:,0].argsort()][::-1]
            f_max2min = sep_sorted[:,1] #extract frequency
            cf = np.cumsum(f_max2min)
            # sep_sorted_with_cf = sep_sorted
            # sep_sorted_with_cf[:,1] = cf
            #interpolate on 'Column 2' and read from 'Column 1'
            if cf[-1] > 1E-4:
                if cf[0] < 1E-4:
                    fl = interp1d(cf,sep_sorted[:,0],kind='linear')
                    sep_dim = fl(1E-4)
                else:
                    # print("Impinged longer than 1 hr ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[0]))    
                    sep_dim = sep_sorted[0,0] #should be 3,600
            else:
                print("Not dimensioning at ({:4.1f},{:4.1f}): {:.2e} < 1E-4/year".format(XX[i],YY[j],cf[-1]))
                sep_dim = 0.            
            ImpingeDuration[i,j,k] = sep_dim


    #smoothed contour
    #https://stackoverflow.com/questions/12274529/how-to-smooth-matplotlib-contour-plot
    # Resample your data grid by a factor of 3 using cubic spline interpolation
    # f_resampled = scipy.ndimage.zoom(f,3)

    contour_levels = [60, 120, 300, 600, 1800, 3600]
    # contour_colors = ['green','lime',  'cyan', 'yellow','olive',  'magenta', 'purple','crimson', 'red']
    contour_colors = ['green','lime',  'cyan', 'yellow','olive', 'magenta']

    ttl = "'Impingement Duration [sec] at Level {:8.1f}".format(zz)
    f = ImpingeDuration[:,:,k].transpose()
    sigma = 0.7
    f_resampled = gaussian_filter(f,sigma)
    fig,ax = plt.subplots(figsize=(8,3))
    img = plt.imread("RubyFPSODeckA.jpg")
    ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])

    # cs1=ax.contourf(XX,YY,f,levels,colors=['yellow','magenta'],alpha=0.7)
    cs1=ax.contourf(XX,YY,f_resampled,contour_levels,colors=contour_colors,alpha=0.7)
    fig.colorbar(cs1)
    cs1.cmap.set_under('w')
    cs1.cmap.set_over('k')
    cs2=ax.contour(XX,YY,f_resampled,contour_levels,colors=('k',),linewidths=(1,))
    ax.clabel(cs2,fmt='%3d',colors='k',fontsize=10)
    ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
    ax.set_aspect('equal')
    ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
    ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
    ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
    ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))

    ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
    ax.yaxis.grid(b=True, linestyle='--',linewidth=2)
    plt.title(ttl)
    # fig.savefig(fn+'v02.png')
    plt.show()
    #All ares can be subject to one-hour long jet fire!


""" f = ImpingeDuration.transpose()
levels = [60, 120, 300, 600, 1800] #Seconds of jet fire impingement duration
fig,ax = plt.subplots()
img = plt.imread("RubyFPSODeckA.jpg")
ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])
cs1=ax.contourf(XX,YY,f,levels,colors=['yellow','magenta'],alpha=0.7)
cs2=ax.contour(XX,YY,f,levels,colors=('k',),linewidths=(3,))
ax.clabel(cs2,fmt='%3d',colors='k',fontsize=14)
ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
ax.set_aspect('equal')
ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
plt.title(ttl)
plt.show() """
