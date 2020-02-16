#Read TVD & Plot it
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import unicodedata
import re
def slugify(value):
    """
    Normalizes string, converts to lowercase, removes non-alpha characters,
    and converts spaces to hyphens.
    """
    # import unicodedata
    # value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore')
    # value = unicode(re.sub('[^\w\s-]', '', value).strip().lower())
    # value = unicode(re.sub('[-\s]+', '-', value))
    value = re.sub('[^\w\s-]', '', value).strip().lower()
    value = re.sub('[-\s]+', '-', value)
    return value
#sfx = '013-01'
for sfx in ['013-01','020-01','020-02','020-03','020-04','021-01',\
    '023-01','023-02', '023-03','023-04',\
    '024-01','024-02','024-03','025-01','025-02','025-03','027-01','027-02','027-03',\
    '045-01','045-02','046-01']:

    tvExl = load_workbook(filename='v08_9m_TV_'+sfx+'.xlsx')
    sh = tvExl['Sheet1']
    new_scn = False
    #iterte for all row
    #read total row

    r = 1
    while r < sh.max_row:
        rv = sh.cell(r,1).value    
        #read the path at the row next toe 'Sceanrio ...'
        if rv != None:
            if "Scenario" in rv:
                #print (r, "in the loop")
                #print(rv)            
                r = r + 1
                path = sh.cell(r,1).value
                r = r + 2
                weather = sh.cell(r,1).value
                weather = weather[9:]
                scn = path[20:]+"-"+weather[9:]
                r = r + 3
                t = sh.cell(r,1).value # read t=0 data
                mm = sh.cell(r,2).value
                rr = sh.cell(r,3).value
                m0 = mm
                r0 = rr
                TVD = [t, mm, rr ]            
                while t != None:
                    mm = sh.cell(r,2).value
                    rr = sh.cell(r,3).value
                    TVD = np.vstack([TVD, [t, mm, rr]])                
                    r = r + 1
                    tp = t
                    t = sh.cell(r,1).value            
                print("Scenario {} ends at time {:8.2f} \n with the remaining mass {:10.2f}( \% of the inventory) and Release rate: {:8.2f})".\
                    format(scn, tp, mm/m0*100, rr))
                #print("End of scenario {}".format(path+weather))        
                #print(r, rv)

                #print out release rate at t=0, t=5min, t=10min, t=30min, t = 60min
                T_gate = [0, 300, 600, 1800, 3500]
                M_gate = [TVD[0,1], 0., 0., 0., 0.]
                RR_gate = [TVD[0,2], 0., 0., 0., 0.]
                Tgi = 1
                tp = 0
                ti = 1
                for t in TVD[1:-2,0]:                
                    if  tp < T_gate[Tgi] and t>=T_gate[Tgi]:
                        M_gate[Tgi] = TVD[ti,1]
                        RR_gate[Tgi] = TVD[ti,2]
                        print("Mass \& Release Rate at {:12.2f} read".format(t))
                    tp = t
                    ti = ti+1
                print("{:12d}{:12d}{:12d}{:12d}{:12d}".format(T_gate[0],T_gate[1],T_gate[2],T_gate[3],T_gate[4]))
                print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(M_gate[0],M_gate[1],M_gate[2],M_gate[3],M_gate[4]))
                print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(RR_gate[0],RR_gate[1],RR_gate[2],RR_gate[3],RR_gate[4]))
                

                #Plot the graph
                #LineType = ('-','--')            
                                
                """ fig,ax1 = plt.subplots()
                
                Time = TVD[:,0]
                Mass = TVD[:,1]
                RelRate = TVD[:,2]

                masscolor = 'tab:blue'
                ax1.set_xlabel('Time [s]')
                ax1.set_ylabel('Mass [kg]',color=masscolor)
                ax1.plot(Time,Mass,color=masscolor)
                ax1.set_ylim(bottom=0)
                ax1.tick_params(axis='y',labelcolor=masscolor)
                
                ax2 = ax1.twinx()
                rrcolor = 'tab:red'
                ax2.set_ylabel('Release Rate [kg/s]',color=rrcolor)
                ax2.plot(Time,RelRate,color=rrcolor)
                ax2.set_ylim(bottom=0)
                ax2.tick_params(axis='y',labelcolor=rrcolor)
                
                plt.title(path[20:])            
                plt.show()
                fn = slugify(path[20:])
                fig.savefig("{}.png".format(fn)) """
        else:
            print("skip line {}".format(r))
        r = r + 1
        #print(r)

