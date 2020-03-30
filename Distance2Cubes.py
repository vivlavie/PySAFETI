#PyDimModeDeck
#Dim jet fire length at 5 minues after the release

from openpyxl import load_workbook
import math
import numpy as np
import dill

def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    # if m>5:
    #     jf = -13.2+54.3*math.log10(m)
    # elif m>0.1:
    #     jf= 3.736*m + 6.
    # else:
    #     jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
def mfit(jl):
    m = np.power(10,math.log10(jl/2.8893)/0.3728) / 55.5
    return m

    
element_dump_filename = 'Bv06_dump'
Area = "ProcessArea"

# element_dump_filename = 'Bv06_offloading_dump'
# Area = "Offloading"

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)

iExl=load_workbook(filename='Cube_DimenScn.xlsx')
shC = iExl['DimCube_0320']

DimCubes = {}

for i in range(3,31):
    J = "J{:02d}".format(shC.cell(i,1).value)
    key = shC.cell(i,2).value
    IS = np.array([shC.cell(i,15).value,shC.cell(i,16).value, shC.cell(i,17).value])
    CC = np.array([shC.cell(i,18).value,shC.cell(i,19).value, shC.cell(i,20).value])
    CP = np.array([shC.cell(i,24).value,shC.cell(i,25).value, shC.cell(i,29).value])
    DimCubes[key] = [IS,CC,CP]
    DC = CC - IS
    lDC = np.sqrt(np.dot(DC,DC))
    if CP[0] != None:
        DP = CP - IS
        lDP = np.sqrt(np.dot(DP,DP))
    for e in lEvent:
        if e.Key == key[:-6]:
            rrc = mfit(lDC/e.jfscale)
            t_e = interp1d(e.TVD[:,2],e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'            
            try:
                tc = t_e(rrc)
            except:
                # print(J, key, " The dimensioning scenario finishes discharge at the time, tc")            
                tc = e.TVD[-1,0]
                rrc = e.TVD[-1,2]
            if CP[0] != None:
                rrp = mfit(lDP/e.jfscale)
                try:
                    tp = t_e(rrp)
                except:
                    # print(J, key, " The dimensioning scenario finishes discharge at the time, tp")            
                    tp = e.TVD[-1,0]
                    rrp = e.TVD[-1,2]
                print("{:4s}{:40s}{:10.1f}{:10.1f}{:10.1f}{:10.1f}{:10.1f}".format(J, key,tc,rrc,tp,rrp,lDP/lDC))
            else:
                print("{:4s}{:40s}{:10.1f}{:10.1f}".format(J,key,tc,rrc))
    

