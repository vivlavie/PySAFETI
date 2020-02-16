from openpyxl import load_workbook
import numpy as np
sfx = '020-01'
tvExl = load_workbook(filename='v08_9m_TV_'+sfx+'.xlsx')
sh = tvExl['Sheet1']
new_scn = False
#iterte for all row
#read total row

r = 1
while r < sh.max_row:
    rv = sh.cell(r,1).value    
    #read the path at the row next toe 'Sceanrio ...'
    if rv != None:
        if "Scenario" in rv:
            #print (r, "in the loop")
            #print(rv)            
            r = r + 1
            path = sh.cell(r,1).value
            r = r + 2
            weather = sh.cell(r,1).value
            weather = weather[9:]
            r = r + 3
            t = sh.cell(r,1).value # read t=0 data
            mm = sh.cell(r,2).value
            rr = sh.cell(r,3).value
            m0 = mm
            r0 = rr
            TVD = [t, mm, rr ]            
            while t != None:
                mm = sh.cell(r,2).value
                rr = sh.cell(r,3).value
                TVD = np.vstack([TVD, [t, mm, rr]])                
                r = r + 1
                tp = t
                t = sh.cell(r,1).value            
            print("Scenario {} ends at time {:8.2f} with the remaining mass {:10.2f}( \% of the inventory) with Release rate: {:8.2f})".\
                format(path+weather, tp, mm/m0*100, rr))
            #print("End of scenario {}".format(path+weather))        
            #print(r, rv)
    else:
        print("skip line {}".format(r))
    r = r + 1
    #print(r)

