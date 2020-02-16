
from openpyxl import load_workbook
import math
import numpy as np
import dill

#NumDirections = 14 # 30 in every horizontal direction & Vertical up & down
#NumDirections = 8 # 60 in every horizontal direction & Vertical up & down
NumDirections = 6 # 6 cones to cover the all release direction for a sphere




#Hull
Area = 'Hull'
iExlFilename='h11i'
cExlFilename='h11c_ExdCrv'
SysLen = 4 #Length of 'h11'

#Separate the closed drain out of HFP
EqvsModule = {'057-01': 'CDA', '043-01':'KOD', '043-02':'KOD', \
    '131-02-fwdP':'HFP', '131-02-fwdS':'HFS', '131-02-aftP':'HAP', '131-02-aftS':'HAS', \
        '131-01-fwdP':'HFP', '131-01-fwdS':'HFS', '131-01-aftP':'HAP', '131-01-aftS':'HAS', \
            '046-01':'MeOHS'}     
SFXFiles = ['043-057','046-131']
TVDprefix = 'h11_TVDischarge_'
element_dump_filename = 'h11_dump_jet'
#Hull Parameters End

#Topside
Area = 'Topside'
iExlFilename='v08_9mi'
cExlFilename='v08_9mc'
SysLen = 7 #Length of 'v08_9m' + 1
EqvsModule = {'020-01': 'S05', '020-02':'S05', \
    '020-03':'S04', '020-04':'S04', \
        '021-01':'S02','023-01':'P03','023-02':'P03', \
            '023-03':'S03', '023-04':'S03', \
                '024-01':'P04', '024-02':'P04','024-03':'P04', '025-01':'P04', '025-02':'P04', \
                    '025-03':'S03','027-01':'P05', '027-02':'P05', '027-03':'P05', \
                        '045-01':'P03','045-02':'P03','046-01':'P02','013-01':'S05'}
SFXFiles = ['013-01','020-01','020-02','020-03','020-04','021-01',\
    '023-01','023-02', '023-03','023-04',\
    '024-01','024-02','024-03','025-01','025-02','025-03','027-01','027-02','027-03',\
    '045-01','045-02','046-01']
TVDprefix = 'v08_9m_TV_'
element_dump_filename = 'v09_dump_jet'
#Topside Parameters End


iExl=load_workbook(filename=iExlFilename+'.xlsx')
cExl=load_workbook(filename=cExlFilename+'.xlsx')


shtTLV = iExl['Time varying leak']
shDc = cExl['Discharge']
shJet = cExl['Jet fire']
shPV = iExl['Pressure vessel']
shDispersion = cExl['Flammable Dispersion']
shFireball = cExl['Fireball']
shExplosion = cExl['Explosions']


def jffit(m):
    if m>5:
        return -13.2+54.3*math.log10(m)
    elif m>0.1:
        return 3.736*m + 6.
    else:
        return 0.


class Event:
    def __init__(self,study_folder_pv,scenario,weather,dc=None,jf=None,epf=None,lpf=None,exp=None,freq=None,x=None,y=None,rh = None, hole = None,PP = None, TT = None,fireball = None, dispersion = None):        
        self.Study_Folder_Pv = study_folder_pv
        self.Path = study_folder_pv + "\\" + scenario
        self.Weather = weather
        self.Key = study_folder_pv + "\\" + scenario + "\\" + weather
        self.Frequency = freq
        self.Discharge = dc
        self.JetFire = jf
        self.EarlyPoolFire = epf
        self.LatePoolFire = lpf
        self.Explosion = exp
        self.Fireball = fireball
        self.Dispersion = dispersion
        self.TVD = []
        self.TVDRead = False
        self.Module = ""
        self.X = x
        self.Y = y
        self.ReleaseHeight = rh
        self.Hole = hole
        self.Holemm = holemm
        self.Pressure = PP
        self.Temperature = TT        
        self.PImdIgn = 1
        self.PDelIgn = 0
        self.PExp_Ign = 1
        self.ESDSuccess = True
        self.BDVSuccess = True
        self.BDVDia = 0
        """     def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # RelDuration = self.TVD[-1,0]
            a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:10s} | {:40s} | {:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} | {:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}*".\
                format(ts,   folder,      pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], self.JetFire.Frequency, self.JetFire.SEP, a, b, c, d,e)
        else:
            fmt = "{:10s} | {:40s} | {:20s} | {:20s} | {:20s} | {:20s} | {:.2e} | {:8.2f} | {:8.2f}".\
                format(ts,   folder,      pv, self.Module, scenario, self.Weather, self.Frequency, self.Discharge.ReleaseRate,self.Discharge.Duration)
        
        return fmt """
    def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # print(self.Path,self.Dispersion,self.JetFire,self.EarlyPoolFire,self.LatePoolFire,self.Explosion,self.Fireball)
            # RelDuration = self.TVD[-1,0]
            # a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} \
                | {:8.2f} [barg]| {:8.2f} [degC]| | ESD Success: {:}| BDV Success {:}\n".\
                format(pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], \
                self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess)
            
            if self.Dispersion != None:
                fmt += "\tDispersion: Frequency {:8.2e}, Distance to LFL {:}\n".format(self.Dispersion.Frequency,self.Dispersion.Distance2LFL)
            else:
                fmt += "\tDispersion: No dispersion\n"
            
            if self.JetFire != None:
                fmt += "\tJF: Frequency {:8.2e}, Length {:6.1f}\n".format(self.JetFire.Frequency,self.JetFire.Length)
            else:
                fmt += "\tJF: No Jet Fire\n"
            
            if self.Explosion != None:
                fmt += "\tExplosion: Frequency {:8.2e}, Overpressure {:6.1f} to {:6.1f} [m]\n".format(self.Explosion.Frequency,self.Explosion.Overpressures[-1],self.Explosion.Distance2OPs[-1])
            else:
                fmt += "\tExplosion: No Explosion\n"
            
            if self.EarlyPoolFire != None:
                fmt += "\tEPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.EarlyPoolFire.Frequency,self.EarlyPoolFire.Diameter)
            else:
                fmt += "\tEPF: No Early Pool Fire\n"
            if self.LatePoolFire != None:
                fmt += "\tLPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.LatePoolFire.Frequency,self.LatePoolFire.Diameter)
            else:
                fmt += "\tLPF: No Late Pool Fire\n"
            if self.Fireball != None:
                fmt += "\tFireball: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.Fireball.Frequency,self.Fireball.Diameter)
            else:
                fmt += "\tFireball: No Fireball\n"
        
        else:
            ts,pv,IS,hole = self.Path.split("\\")
            f = "{:15s} {:4s} ({:6.1f},{:6.1f},{:6.1f}) {:6.1f} {:6s} P({:8.2f}) T({:8.2f}) ESD({:}) BDV({:}) ".\
                        format(IS, self.Module, self.X, self.Y,  self.ReleaseHeight, self.Holemm, self.Weather[9:],self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess)
            if self.Dispersion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.JetFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.EarlyPoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.LatePoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Explosion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Fireball != None:
                f += "  O  "
            else:
                f += "  X  "            
            fmt = f
        
        return fmt
        # return self.Key
    def EventTree(self):
        # Pwss = 1/8*1/2 #8 wind directions & 2 wind velocities
        Pwss = 1
        Pi = self.PImdIgn # Probability of immediate ignition #input into SAFETI hole        
        Pdi = self.PDelIgn

        Ph = 4/6 #horizontal fraction for vertical and horizontal release cases
        Pj_h = 1/4*Ph # probability of horizontal jet fire Pjh x ph = 4/6 x 1/4 = 1/6 with no accompanying pool fire
        Pp_h = 1/2*Ph #probability of pool fire upon horizontal release with no accompanying jet fire
        Pjp_h =0.5/4*Ph #probability of jet & pool fire upon horizontal release

        Pj_v = 1/2*(1-Ph)#probabilty of vertical jet fire upon vertical release
        Pp_v = 1/2*(1-Ph)#
        Pjp_v = 0.5/4*(1-Ph)#

        #short release
        Ps = 1 #fraction for short release effects; release is shorter than cut-off time '20 s'
        Pbp = 1 #prob of bleve & pool fire
        Pb = 0 #prob of bleve without a pool fire
        Pfp = 0 #prob for short duration release immediate ignition resulting in flash fire and pool
        Pf = 0 #Flash fire along
        Pp = 0#prob for short duration release immediate ignition resulting in  to have pool fire
        Pep = 0#prob for short duration release immediate ignition resulting in  to have immediate explosion with pool fire
        Pe = 0#prob for short duration release immediate ignition resulting in  to have immediate explosion without pool fire
        Pp = 0#prob for short duration release immediate ignition resulting in  to have early pool fire without explosion


        Prp = 0.15 #prob of residual pool fire
        Po = (1-Pi)*Pwss*Pipr*Prp#prob of igniton of residual pool
        Pf_d = 0.6 #prob of delayed flash fire
        Pe_d = 0.4 #prob of explosion upon delayed ignition

        #Immediate, Not short
        P_i_ns_h_j = (1-Ps)*Ph*Pwss*Pj_h*Pi
        P_i_ns_h_p = (1-Ps)*Ph*Pwss*Pp_h*Pi
        P_i_ns_h_jp = (1-Ps)*Ph*Pwss*Pjp_h*Pi
        P_i_ns_v_j = (1-Ps)*(1-Ph)*Pwss*Pj_v*Pi
        P_i_ns_v_p = (1-Ps)*(1-Ph)*Pwss*Pp_v*Pi
        P_i_ns_v_jp = (1-Ps)*(1-Ph)*Pwss*Pjp_v*Pi
        #Immediate, Short
        P_i_s_BP = Ps*Pwss*Pbp*Pi #BLEVE & Pool
        P_i_s_B = Ps*Pwss*Pb*Pi #BLEVE only
        P_i_s_FP = Ps*Pwss*Pfp*Pi #Flash fire and pool
        P_i_s_F = Ps*Pwss*Pf*Pi #Flash fire only
        P_i_s_EP = Ps*Pwss*Pep*Pi #Explosion & pool fire
        P_i_s_E = Ps*Pwss*Pe*Pi #Explosion only
        P_i_s_P = Ps*Pwss*Pp*Pi #Pool fire only

        #Delayed
        P_d_FP = (1-Pi)*Pwss*Pfp*Pdi #Delayed flash and pool fire
        P_d_E = (1-Pi)*Pwss*Pdd*Pdi #Delayed explosion
        P_d_p = (1-Pi)*Pwss*Prp*Pirp #Redisual pool fire

        print("----Imd Ignition---Not short--|-Horizontal---Jet fire       : {:8.2e}".format(P_i_ns_h_j))
        print("                 |             |             |-Pool fire      : {:8.2e}".format(P_i_ns_h_p))
        print("                 |             |             -Jet & Pool fire: {:8.2e}".format(P_i_ns_h_jp))
        print("                 |             |-Vertical-----Jet fire       : {:8.2e}".format(P_i_ns_v_j))
        print("                 |                           |-Pool fire      : {:8.2e}".format(P_i_ns_v_p))
        print("                 |                           -Jet & Pool fire: {:8.2e}".format(P_i_ns_v_jp))
        print("                 --Short----------------------BLEVE & Poolfire: {:8.2e}".format(P_i_s_BP))
        print("                                            |-BLEVE only      : {:8.2e}".format(P_i_s_B))
        print("                                            |-Flash & Pool fire:{:8.2e}".format(P_i_s_FP))
        print("                                            |-Flash only: {:8.2e}".format(P_i_s_F))
        print("----Delayed-------Ignited--------------------Flash and Pool fire:{:8.2e}".format(P_d_FP))
        print("                                            |- Explosion : {:8.2ef}".format(P_d_E))
        print("                                            |- Pool fire: {:8.2e}".format(P_d_p))
        

class JetFire:
    def __init__(self,length,sep,jff):
        self.Length = length
        self.SEP = sep
        self.Frequency = jff
        self.Ts = [0.,0.,0.,0.,0.]
        self.JetLengths = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.JetLengths
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)        
        return fmt

class Explosion:
    def __init__(self,fexp,op,d2op):
        self.Overpressures = op
        self.Distance2OPs = d2op
        self.Frequency = fexp        
    def __str__(self):
        a,b,c = self.Distance2OPs
        pa,pb,pc = self.Overpressures
        fmt = "{:.2e} | {:8.2f}barg to {:8.2f} | {:8.2f}bar to {:8.2f} | {:8.2f} barg to {:8.2f}".\
            format(self.Frequency, pa, a, pb, b, pc,c)        
        return fmt

class Dispersion:
    def __init__(self,fdisp,d2f,d):
        self.Distance2LFLFraction = d2f
        self.Distance2LFL = d
        self.Frequency = fdisp        
    def __str__(self):
        
        fmt = "{:.2e} | {:8.2f}m to LFL Fraction | {:8.2f}m to LFL".\
            format(self.Frequency, self.Distance2LFLFraction, self.Distance2LFL)        
        return fmt
class Fireball:
    def __init__(self,ffb,D,sep):
        self.Diameter = D
        self.SEP = sep
        self.Frequency = ffb       
    def __str__(self):
        
        fmt = "{:.2e} | Diameter: {:8.2f} | SEP {:8.2f}".\
            format(self.Frequency, self.Diameter, self.SEP)        
        return fmt

class Discharge:
    def __init__(self,rr,duration):
        self.ReleaseRate = rr
        self.Duration = duration
        self.Ts = [0.,0.,0.,0.,0.]
        self.Ms = [0.,0.,0.,0.,0.]
        self.RRs = [0.,0.,0.,0.,0.]        
    
class EarlyPoolFire:
    def __init__(self,diameter,sep,epff):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

class LatePoolFire:
    def __init__(self,diameter,sep,epff):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

#Read Leaks
#listLeak = []
ReleaseHeight= {}
Frequency ={}
PImdIgn ={}
PDelIgn = {}
PExp ={}
ReleaseRate = {}
Duration = {}
J00 = {}
J05 = {}
J10 = {}
J30 = {}
J60 = {}
JetSEP = {}
JetFrequency = {}
Tiso = {}
Tbdv = {}
Hole = {}
Holemm = {}
PP = {} #Pressure of the PV
TT  = {} #Temp of the PV
ESDSuccess ={}
BDVSuccess ={}
BDVDia ={}
X = {}
Y = {}

#for r in range(6,numscn):
r = 63
while shtTLV.cell(r,1).value == "Yes":
    study = shtTLV.cell(r,2).value
    folder = shtTLV.cell(r,3).value
    equip  = shtTLV.cell(r,15).value
    hole = shtTLV.cell(r,27).value    
    holemm = shtTLV.cell(r,28).value    
    key = study + "\\" + folder + "\\" + equip + "\\" + hole
    relheight = shtTLV.cell(r,32).value

    #Construci disctionarys with keys as 'Path'
    Hole[key] = hole
    Holemm[key] = holemm
    Frequency[key]  = shtTLV.cell(r,38).value
    PImdIgn[key] = shtTLV.cell(r,44).value
    PDelIgn[key] = shtTLV.cell(r,46).value
    PExp[key] = shtTLV.cell(r,48).value

    #Safety performance parameters for a hole
    Tiso[key] = shtTLV.cell(r,64).value
    Tbdv[key] = shtTLV.cell(r,67).value
    ESDSuccess[key] = shtTLV.cell(r,63).value
    BDVSuccess[key] = shtTLV.cell(r,66).value
    BDVDia[key] = shtTLV.cell(r,68).value
    r=r+1
    print(r)
numLeak = r-63



lEvent = [] #List of class Event
#Construct Event list
#Read Discharge, shDc
weathers = ['Category 1.2/F', 'Category 6.6/D']
k=0
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shDc.max_row+1):
    study_folder_pv = shDc.cell(r,1).value
    # study,folder,pv = study_folder_pv.split("\\")
    scenario = shDc.cell(r,2).value
    weather = shDc.cell(r,3).value
    path = study_folder_pv + "\\" + scenario
    key = study_folder_pv + "\\" + scenario + "\\" + weather
    if weather in weathers:
        rr = shDc.cell(r,10).value
        ReleaseRate[key] = rr
        duration = shDc.cell(r,11).value
        Duration[key] = duration        
        aEvent = Event(study_folder_pv,scenario,weather,freq=Frequency[path],dc=Discharge(rr,duration),hole=Hole[path])
        lEvent.append(aEvent)
    else:
        print("{} Discharge events read".format(k))
        break
    k=k+1


for e in lEvent:    
    study,folder,pv,hole,weather = e.Key.split("\\")
    path = study + "\\" + folder + "\\" + pv + "\\" + hole
    e.PImdIgn = PImdIgn[path]
    e.PDelIgn = PDelIgn[path]
    e.PExp_Ign = PExp[path]
    # e.ESDSuccess = 
    # e.BDVSuccess = 

#for r in range(0,numLeak*len(weathers)):
# for r in range(0,shDc.max_row-1):
#     print(lEvent[r])
# print (k)

# for e in lEvent:
#     print(e.Key)

#Read TV Discharge data and append the info on each Event ... equip - hole - weather

import numpy as np
#import matplotlib.pyplot as plt
for sfx in SFXFiles:
#for sfx in ['013-01']:
    #sfx = '020-01'
    tvExl = load_workbook(filename=TVDprefix + sfx+'.xlsx')
    
    sh = tvExl['Sheet1']
    new_scn = False
    #iterte for all row
    #read total row

    r = 1
    while r < sh.max_row:
        rv = sh.cell(r,1).value    
        #read the path at the row next toe 'Sceanrio ...'
        if rv != None:
            if "Scenario" in rv:
                #print (r, "in the loop")
                #print(rv)            
                r = r + 1
                path = sh.cell(r,1).value
                path = path[SysLen:]
                r = r + 2
                weather = sh.cell(r,1).value
                weather = weather[9:]
                #scn = path[20:]+"-"+weather[9:]
                scn = path +"\\"+weather
                r = r + 3
                t = sh.cell(r,1).value # read t=0 data
                mm = sh.cell(r,2).value
                rr = sh.cell(r,3).value
                m0 = mm
                r0 = rr
                TVD = [t, mm, rr ]            
                
                while t != None:
                    mm = sh.cell(r,2).value
                    rr = sh.cell(r,3).value
                    TVD = np.vstack([TVD, [t, mm, rr]])  
                    if (("Flow line" in path) and ("NE_TLV" in path) and (t==3600)):
                        print([t, mm, rr])
                        print(TVD[-1,:])

                    tp = t
                    r = r + 1
                    t = sh.cell(r,1).value    
                
                
                if tp==3600:
                    print(path, weather, TVD[-1,:])
                #print("Scenario {} ends at time {:8.2f} \n with the remaining mass {:10.2f}( \% of the inventory) and Release rate: {:8.2f})".\
                #    format(scn, tp, mm/m0*100, rr))
                #print("End of scenario {}".format(path+weather))        
                #print(r, rv)
                


                #print out release rate at t=0, t=5min, t=10min, t=30min, t = 60min
                T_gate = [0, 300, 600, 1800, 3600]
                M_gate = [TVD[0,1], 0., 0., 0., 0.]
                RR_gate = [TVD[0,2], 0., 0., 0., 0.]
                #Tgi = 1
                for Tgi in range(1,5):
                    tp = 0
                    ti = 1
                    #if T_gate[Tgi] < Duration[scn]:
                    for t in TVD[1:,0]:                
                        if  tp < T_gate[Tgi] and T_gate[Tgi] <= t:
                            #t = TVD[ti,0]
                            M_gate[Tgi] = TVD[ti,1]
                            RR_gate[Tgi] = TVD[ti,2]
                            #print("ti={:5d} Mass {:12.2f} ({:12.2f}) \& Release Rate {:12.2f} ({:12.2f})at {:12.2f} read".format(ti,M_gate[Tgi],TVD[ti,1],RR_gate[Tgi],TVD[ti,2],t))
                        tp = t
                        ti = ti+1
                #print("{:12d}{:12d}{:12d}{:12d}{:12d}".format(T_gate[0],T_gate[1],T_gate[2],T_gate[3],T_gate[4]))
                #print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(M_gate[0],M_gate[1],M_gate[2],M_gate[3],M_gate[4]))
                #print("{:12.2f}{:12.2f}{:12.2f}{:12.2f}{:12.2f}".format(RR_gate[0],RR_gate[1],RR_gate[2],RR_gate[3],RR_gate[4]))

                #Find an Event that fits the 'path' + 'weather'
                i=0
                while (not (lEvent[i].Path == path and lEvent[i].Weather == weather)):
                    i = i + 1
                    if i == len(lEvent):
                        break
                if i < len(lEvent):
                    key = path + "\\" + weather
                    #Save the results of T, M & RR
                    lEvent[i].Discharge.Ts = T_gate
                    lEvent[i].Discharge.Ms = M_gate
                    lEvent[i].Discharge.RRs = RR_gate
                    lEvent[i].TVDRead = True
                    lEvent[i].TVD = TVD
                else:
                    print("Event corresponding to {} {} not found".format(path, weather))
                #Copy the discharge pattern from 6.6/D to 1.2/F
                i=0
                while (not (lEvent[i].Path == path and lEvent[i].Weather == 'Category 1.2/F')):
                    i = i + 1
                    if i == len(lEvent):
                        break
                if i < len(lEvent):
                    key = path + "\\" + 'Category 1.2/F'
                    lEvent[i].Discharge.Ts = T_gate
                    lEvent[i].Discharge.Ms = M_gate
                    lEvent[i].Discharge.RRs = RR_gate
                    lEvent[i].TVDRead = True
                    lEvent[i].TVD = TVD
                    #print(lEvent[i].Discharge.Ms)
                else:
                    print("Event corresponding to {} {} not found".format(path, 'Category 1.2/F'))

                
        else:
            print("skip line {}".format(r))
        r = r + 1
#End of reading Discharge files

#For an Event, find the relevant Leak and return its frequency
#r = 10
#key = lEvent[r].Path
#print("Path: {}  Frequency: {} Release Height: {}".format(key,Frequency[key],ReleaseHeight[key]))

#Reading Jet fire
#1. Find an Event that has the same study_folder_pv, scenario, & weather
# lEvent[##].JetFire = JetFiire
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shJet.max_row+1):
    study_folder_pv = shJet.cell(r,1).value
    scenario = shJet.cell(r,2).value
    weather = shJet.cell(r,3).value
    path = study_folder_pv + "\\" + scenario
    i=0
    while not (lEvent[i].Path == path and lEvent[i].Weather == weather):
        i = i + 1
        if i == len(lEvent):
            break
    if i < len(lEvent):
        key = path + "\\" + weather
        if weather in weathers:
            jfl = shJet.cell(r,10).value
            #J00[key] = jfl
            sep = shJet.cell(r,11).value
            JetSEP[key] = sep

            a_event = lEvent[i]
            jff = Frequency[path]*PImdIgn[path]/NumDirections
            if weather == 'Category 1.2/F':
                jff = jff*0.18
            elif weather == 'Category 6.6/D':
                jff = jff*0.82
            JetFrequency[key] = jff
            lEvent[i].JetFire = JetFire(jfl,sep,jff)

            if a_event.Discharge.Ts[2] != 0:
                jf0 = a_event.JetFire.Length
                m0 = a_event.Discharge.Ms[0]
                jf0fit = jffit(m0)
                jfscale = a_event.JetFire.Length/jf0fit
                lEvent[i].JetFire.JetLengths[0] = jf0
                for ti in range(1,5):
                    #t = a_event.Ts[i]
                    m = a_event.Discharge.Ms[ti]
                    jf = jffit(m)
                    lEvent[i].JetFire.JetLengths[ti] = jf*jfscale
                    #lEvent[i].JetFire.JetLengths[ti] = jf
                J00[key] = jf0
                J05[key] = lEvent[i].JetFire.JetLengths[1]
                J10[key] = lEvent[i].JetFire.JetLengths[2]
                J30[key] = lEvent[i].JetFire.JetLengths[3]
                J60[key] = lEvent[i].JetFire.JetLengths[4]

        else:
            print("{} Jet fire events read".format(k))
            break
    k=k+1
#Read Jet, shJet


r=63
# while shPV.cell(r,1).value == "Yes":
#     equip  = shPV.cell(r,15).value
#     X[equip] = shPV.cell(r,152).value
#     Y[equip] = shPV.cell(r,153).value
#     ReleaseHeight[equip] = shPV.cell(r,38).value
#     r += 1
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    folder  = shPV.cell(r,3).value
    pv  = shPV.cell(r,15).value
    key = study  + "\\" +  folder  + "\\" + pv
    X[key] = shPV.cell(r,152).value
    Y[key] = shPV.cell(r,153).value
    ReleaseHeight[key] = shPV.cell(r,38).value
    PP[key] = shPV.cell(r,24).value
    TT[key] = shPV.cell(r,23).value
    r += 1
#Read X&Y and Put into lEvent
for e in lEvent:
    e.X = X[e.Study_Folder_Pv]
    e.Y = Y[e.Study_Folder_Pv]
    e.ReleaseHeight = ReleaseHeight[e.Study_Folder_Pv]
    e.Pressure = PP[e.Study_Folder_Pv]
    e.Temperature = TT[e.Study_Folder_Pv]
    e.Holemm = Holemm[e.Path]
    # print(e.X, e.Y, e.ReleaseHeight)
for k in EqvsModule.keys():
     for e in lEvent:         
         if k in e.Key:
             e.Module = EqvsModule.get(k)

shPool = cExl['Early Pool Fire']
EPSEP = {}
EPFreq = {}
#Read Early Pool Fire
for r in range(2,shPool.max_row+1):
    study_folder_pv = shPool.cell(r,1).value
    scenario = shPool.cell(r,2).value
    weather = shPool.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    key = path + "\\" + weather    
    sep = shPool.cell(r,11).value    
    EPSEP[key] = sep
    dia = shPool.cell(r,10).value    
    for e in lEvent:
        if e.Key == key:
            epff = e.JetFire.Frequency
            EPFreq[key] = epff
            e.EarlyPoolFire = EarlyPoolFire(dia,sep,epff)





shPool = cExl['Late Pool Fire']
LPSEP = {}
LPFreq = {}
#Read Early Pool Fire
for r in range(2,shPool.max_row+1):
    study_folder_pv = shPool.cell(r,1).value
    scenario = shPool.cell(r,2).value
    weather = shPool.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    key = path + "\\" + weather    
    sep = shPool.cell(r,11).value    
    LPSEP[key] = sep
    dia = shPool.cell(r,10).value    
    for e in lEvent:
        if e.Key == key:
            epff = e.JetFire.Frequency
            LPFreq[key] = epff
            e.EarlyPoolFire = EarlyPoolFire(dia,sep,epff)

#Pool fire total frequency

#shExplosion
ExpD1 = {}
ExpD2 = {}
ExpD3 = {}
ExpFreq = {}
#Read Early Pool Fire
for r in range(2,shExplosion.max_row+1):
    study_folder_pv = shExplosion.cell(r,1).value
    scenario = shExplosion.cell(r,2).value
    weather = shExplosion.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    key = path + "\\" + weather    
    p1 = shExplosion.cell(r,9).value    
    p2 = shExplosion.cell(r,10).value    
    p3 = shExplosion.cell(r,11).value    
    d1 = shExplosion.cell(r,12).value    
    d2 = shExplosion.cell(r,13).value    
    d3 = shExplosion.cell(r,14).value    

    ExpD1[key] = d1
    ExpD2[key] = d2
    ExpD3[key] = d3
    
    for e in lEvent:
        if e.Key == key:
            fexp = e.JetFire.Frequency/e.PImdIgn*e.PDelIgn*e.PExp_Ign
            ExpFreq[key] = fexp
            e.Explosion = Explosion(fexp,[p1,p2,p3],[d1,d2,d3])


""" class Dispersion:
    def __init__(self,fdisp,d2f,d):
        self.Distance2LFLFraction = d2f
        self.Distance2LFL = d
        self.Frequency = fdisp """


#shDispersion
DistLFLFrac = {}
DistLFL = {}
DispFreq = {}
#Read Flammable Dispersion
for r in range(2,shDispersion.max_row+1):
    study_folder_pv = shDispersion.cell(r,1).value
    scenario = shDispersion.cell(r,2).value
    weather = shDispersion.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    key = path + "\\" + weather    
    
    dfrac = shDispersion.cell(r,12).value    
    d = shDispersion.cell(r,13).value    
    
    DistLFLFrac[key] = dfrac
    DistLFL[key] = d
    
    for e in lEvent:
        if e.Key == key:
            fdisp = e.JetFire.Frequency
            DispFreq[key] = fdisp
            e.Dispersion = Dispersion(fdisp,dfrac,d)

#shFireball
DiaFireball = {}
SEPFireball = {}
FireballFreq = {}
#Read Flammable Dispersion
for r in range(2,shFireball.max_row+1):
    study_folder_pv = shFireball.cell(r,1).value
    scenario = shFireball.cell(r,2).value
    weather = shFireball.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    key = path + "\\" + weather    
    
    dia = shFireball.cell(r,9).value    
    sep = shFireball.cell(r,10).value    
    
    DiaFireball[key] = dia
    SEPFireball[key] = sep
    
    for e in lEvent:
        if e.Key == key:
            ffireball = e.JetFire.Frequency
            FireballFreq[key] = ffireball
            e.Fireball = Fireball(ffireball,dia,sep)


with open(element_dump_filename,'wb') as element_dump:
    dill.dump(lEvent,element_dump)

F = 0.
for e in lEvent:
    # if e.Discharge.ReleaseRate > 0.1:
    if e.JetFire.SEP > 50:
        F += e.JetFire.Frequency
    else:
        print(e.Key + "is excluded from Jet fire frequency summation")
print("{:15s} Jet fire Toal Frequency: {:.2e}".format(Area, F))
print("{:15s} Pool fire Toal Frequency: {:.2e}".format(Area, sum(EPFreq.values()) + sum(LPFreq.values())))
