#PyDimCube
#For a cube, read from each cube the duration that jet fire impinges on it
#PyDimCube3
#To consider failure cases for fire or gas detector
#Failure of fire detector for immediate ignition
#Failure of gas detectin for delayed ignition

#PyDimCube4 using 'Bv06_c.xlsx'


from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill
import sys

import matplotlib.pylab as pltparam


# sys.stdout = open('DimCube2.txt','w')

FireWallBtn3and4 = True
WantPlot = False
cube_result2file = False
FirewallX = 140.7 #meter from AP

def DtoH (D05):
    if D05 > 19.7:
            H05 = 0.4664*D05 + 18.345
    elif D05 < 1.0:
        H05 = 0.
    else: #  1.0 < D05 < 19.7:
        H05 = 1.2592*D05 + 2.7235
    return H05
def PD(t,De,Te):
    if t < Te:
        PD = De * (1-math.sqrt(t/Te))
    else:
        PD = 0
    return PD
#pool burning rate 0.062 kg/m2-s
#MS ; spilt mass
#Pool fire duration
# Ms / (3.14*D^2/4 * br) *3/2 where 3/2 is a factor to consider shringking pool

def print_cum_cube(cube,AA):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube))
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F))
        if F>1.0E-3:
            break
def print_cum_cube_file(cube,AA,a_file):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube),file = a_file)
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F),file = a_file)
        if F>1.0E-3:
            break


element_dump_filename = 'Bv06_dump'
icubeloc='SCE_CUBE_XYZ2_Process'

# element_dump_filename = 'Bv06_hull_dump'
# icubeloc='SCE_CUBE_XYZ2_HullDeck'

# element_dump_filename = 'Bv06_offloading_dump'
# icubeloc='SCE_CUBE_XYZ2_Offloading'

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)

#Read 'Presure vessel' from Input
iExlFilename='Bv06_i'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
Z = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv    
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    r += 1
numPV = r-63
def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    if m>5:
        jf = -13.2+54.3*math.log10(m)
    elif m>0.1:
        jf= 3.736*m + 6.
    else:
        jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
iIS=load_workbook(filename='IS_v12_shk.xlsx')
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 82:
    pv = shIS.cell(r,5).value
    IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    Modules[pv] = shIS.cell(r,7).value
    Deck[pv] =  shIS.cell(r,8).value
    if shIS.cell(r,24).value != None:
       nedvs = shIS.cell(r,24).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1



# iExlFilename='025-02-01-G_i'
# iExl=load_workbook(filename=iExlFilename+'.xlsx')
# cExlFilename='025-02-01-G_c'

#Failure of fire detector for immediate ignition
P_FD_Fail = 0.05 #default value? #should be adjusted for hole size or leak rate? To be suggested to use the release rate!
# P_FD_Fail['SM'] = 0.1
# P_FD_Fail['ME'] = 0.05
# P_FD_Fail['MA'] = 0.005
# P_FD_Fail['LA'] = 0.005

#Failure of gas detectin for delayed ignition
P_GD_Fail = 0.05




iExl=load_workbook(filename=icubeloc+'.xlsx')
shCube = iExl['xyz']
ncube = shCube.cell(1,1).value
Xcube = {}
Ycube = {}
Zcube = {}
Cubes = []
CubeDeck = {}
CubeImpingeDuration = {}
for i in range(3,ncube+3):
    id = shCube.cell(i,1).value
    Cubes.append(id)
    Xcube[id] = shCube.cell(i,2).value
    Ycube[id] = shCube.cell(i,3).value
    Zcube[id] = shCube.cell(i,4).value
    CubeDeck[id] = shCube.cell(i,5).value

print("{:32s} {:5s} {:4s} {:10s} Duration[sec]  Duration[min] m_dot[kg/s] or Pool diameter[m]".\
                format('Scenario','Module','Deck','Cube'))
ImpingeDurationArray = []

for i in range(0,ncube):    
# for i in range(0,37):
# for i in [ncube-1]:
    # li = 0
    # sep = np.zeros((len(lEvent),2))
    ImpingeDurationArray = []
    id = Cubes[i]
    xx = Xcube[id]
    yy = Ycube[id]
    zz = Zcube[id]
    if cube_result2file == True:
        f_cube_result = open(id+".txt","w")
    # if (zz >= 35.0) and (zz < 44):
    #     CubeDeck = 'A'
    # elif (zz >= 44.0) and (zz < 53):
    #     CubeDeck = 'B'
    # elif (zz >= 53):
    #     CubeDeck = 'C'
    # elif (zz < 35 ):
    #     CubeDeck = 'Hull'
    
    for ei in range(0,len(lEvent)):
        e = lEvent[ei]
        if (CubeDeck[id] == e.Deck) or (id[0:3]) == e.Module:
            #Exposure to jet fire
            #read frequency
            # sep[li,1] = e.JetFire.Frequency
            #check if the receptor is in the flame
            dx = xx - e.X
            dy = yy - e.Y
            dz = zz - (e.ReleaseHeight+35)
            rr = math.sqrt(dx*dx+dy*dy+dz*dz)
            #Array of jet length
            jl_e = e.jfscale*2.8893*np.power(55.5*e.TVD[:,2],0.3728) #the 3rd column of the matrix, Lowesmith formulat
            t_e = interp1d(jl_e,e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'            

            if e.Discharge.ReleaseRate <= 1:
                P_FD_Fail = 0.1
                P_GD_Fail = 0.1
            elif e.Discharge.ReleaseRate <= 10:
                P_FD_Fail = 0.05
                P_GD_Fail = 0.05
            elif e.Discharge.ReleaseRate > 10:
                P_FD_Fail = 0.005
                P_GD_Fail = 0.005
            else:
                print('something wrong in P_FD_Fail')

            # The following condition is to consider effects of fire wall
            # If the firewall is considered and the source and targets are at different areas, we skip considering the jet fire effect.
            # In other words, if the fire is not considered (firewall == false) or (if they are in the same area), we will consider the jet fire
            if (FireWallBtn3and4 == False) or (((xx > FirewallX) and (e.X > FirewallX)) or ((xx < FirewallX) and (e.X < FirewallX))):
                if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                    #Fire detection successful
                    ff = e.JetFire.Frequency*(1-P_FD_Fail)
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FO_Jet"])                    

                    #Fire detection failure, any release irrespective of success of ESDV and BDV will result in EXBX release case      
                    ff = e.JetFire.Frequency/e.PESD/e.PBDV*P_FD_Fail
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FX_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FX_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FX_Jet"])                    
                else:  #Fire detection succedded
                    ff = e.JetFire.Frequency*(1-P_FD_Fail)
                    if max(jl_e) < rr:                                     
                        ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                    elif min(jl_e) > rr:                
                        ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                    else:                
                        ImpingeDurationArray.append([t_e(rr),ff,e.Key+"FO_Jet"])                    
            

            if (id[0:3]) == e.Module:
                #Exposure to pool fire
                if e.EarlyPoolFire != None:
                    #Read pool fire duration, ho
                    dd = e.EarlyPoolFire.Diameter
                    Ms = e.Discharge.Ms[0] - e.Discharge.Ms[-1]
                    PFD = Ms/(3.14*dd*dd/4 * 0.062)*3/2 #Pool Fire Duration
                    rr = math.sqrt(dx*dx+dy*dy)

                    if (rr < dd) & (zz > e.ReleaseHeight) & (PFD > 0):
                        di = PFD*(dd-rr)/dd                        

                        if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                            #Fire detection successful
                            ff = e.EarlyPoolFire.Frequency*(1-P_FD_Fail)
                            ImpingeDurationArray.append([di,ff,e.Key+"FO_EarlyPool"])                        
                        #Fire detection failure
                            ff = 0.
                            pv,hole,weather = e.Key.split("\\")
                            for ee in lEvent:
                                if ((pv in ee.Key) and (hole[:2] == ee.Hole[:2]) and (weather == ee.Weather)):
                                    ff += ee.EarlyPoolFire.Frequency*P_FD_Fail
                            ImpingeDurationArray.append([di,ff,e.Key+"FX_EarlyPool"])                                                    
                        else:                          
                            ImpingeDurationArray.append([di,e.EarlyPoolFire.Frequency*(1-P_FD_Fail),e.Key+"FO_EarlyPool"])
                if e.LatePoolFire != None:
                    #Read pool fire duration, ho
                    dd = e.LatePoolFire.Diameter
                    Ms = e.Discharge.Ms[0] - e.Discharge.Ms[-1]
                    PFD = Ms/(3.14*dd*dd/4 * 0.062)*3/2 #Pool Fire Duration
                    rr = math.sqrt(dx*dx+dy*dy)
                    if (rr < dd) & (zz > e.ReleaseHeight) & (PFD > 0):
                        di = PFD*(dd-rr)/dd                        
                        if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                            #Fire detection successful
                            ff = e.LatePoolFire.Frequency*(1-P_GD_Fail)
                            ImpingeDurationArray.append([di,ff,e.Key+"FO_LatePool"])                        
                        #Fire detection failure
                            ff = 0.
                            pv,hole,weather = e.Key.split("\\")
                            for ee in lEvent:
                                if ((pv in ee.Key) and (hole[:2] == ee.Hole[:2]) and (weather == ee.Weather)):
                                    ff += ee.LatePoolFire.Frequency*P_GD_Fail
                            ImpingeDurationArray.append([di,ff,e.Key+"FX_LatePool"])                                                    
                        else:                          
                            ImpingeDurationArray.append([di,e.LatePoolFire.Frequency*(1-P_GD_Fail),e.Key+"FO_LatePool"])
      
    #To pin-point a scenario that give the dimensioning scenario
    IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0]) #with the longest duration at the bottom
    
    cf = 0
    jp = IDAsorted[-1]
    DimFreq = 1.0E-4
    di = 0
    cfp = jp[1]#frequency for the longest duration jet fire
    InterpolationSuccess = False
    if cfp > DimFreq:
        di = jp[0]
        scn = jp[2]
    else:
        for j in IDAsorted[-2:0:-1]:
            cf = cfp + j[1]
            if cf >= DimFreq and cfp < DimFreq:
                dp = jp[0]
                dn = j[0]
                # di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
                # print('Dimensioning jet duration {:8.1f}'.format(di))
                scnp = jp[2]
                scn = j[2]
                InterpolationSuccess = True
                #Choose the scenario close to the threshold
                if abs(cf-DimFreq) > abs(cfp-DimFreq):
                    scn = scnp
                    di = dp
                else:
                    di = dn
                break
            cfp = cf
            # print(cfp,cf)
            jp = j
    
    # print(scn)
    if InterpolationSuccess:
        CubeImpingeDuration[Cubes[i]] = di
        if "Pool" in scn:
            for e in lEvent:
                if e.Key in scn:
                    break
            pv,hole,weather = scn.split("\\")
            weather,pl = weather.split("_")
            if "Early" in pl:
                PFD = e.EarlyPoolFire.Diameter
            elif "Late" in pl:
                PFD = e.LatePoolFire.Diameter
            else:
                print("343: something wrong")
            print("IS: {:20s}, Cube: {:10s} Duration: {:8.1f} Pool diameter: {:8.1f}".format(scn,id,di,PFD))
            # print("{:35s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} ".format(scn, Modules[pv],Deck[pv],id,di,di/60,PFD))
            if cube_result2file == True:
                print("{:35s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} ".\
                    format(scn, Modules[pv],Deck[pv],id,di,di/60,PFD),file=f_cube_result)
                print_cum_cube_file(id,IDAsorted,f_cube_result) 

        else:
            pv,hole,weather = scn.split("\\")            
            weather_fire = weather            
            x1 = X[pv]
            y1 = Y[pv]
            z1 = Z[pv]
            for e in lEvent:
                if e.Key in scn:
                    break
            rrifound = False
            rri = 0
            rr5min = 0.
            for i in range(0,len(e.TVD)-1):
                if (e.TVD[i,0] < di) and (e.TVD[i+1,0] >= di):
                    rri = e.TVD[i,2]
                    rr5min = e.Discharge.RRs[1]
                    rrifound = True
                    break
            if rrifound == False:
                print(id,"Something wrong, rri not found", di, e.TVD[-1,0])

            z1 = z1+35
            dx=xx-x1
            dy=yy-y1
            dz=zz-z1
            jli = jffit(rri)
            xm = x1 + 0.61*dx
            ym = y1 + 0.61*dy
            zm = z1 + 0.61*dz
            # ll = math.sqrt(dx*dx+dy*dy+dz*dz)
            # print("IS: {:35s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:8.1f}{:8.1f}".\
            #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,di,rri))
            print("{:35s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} {:8.1f}".\
                format(scn, Modules[pv],Deck[pv],id,di,di/60,rri,rr5min))
            if cube_result2file == True:                
                print_cum_cube_file(id,IDAsorted,f_cube_result) 

            # print("IS: {:35s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:4s}{:2s} -> {:8s}{:8.1f}{:8.1f}{:8.1f}".\
            #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,Modules[pv],Deck[pv],id,di,rri,2.8893*np.power(55.5*rri,0.3728)))
            # print("IS: {:35s} {:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Release rate: {:8.1f}".\
                # format(scn,x1,y1,z1,id,xx,yy,zz,rri))
            # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*x1,1000*y1,1000*z1,1000*xm,1000*ym,1000*zm,0,0.12*jli*1000))
            # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*xm,1000*ym,1000*zm,1000*xx,1000*yy,1000*zz,0.12*jli*1000,0))
            # print_cum_cube(id,IDAsorted) 
    elif not InterpolationSuccess:
        print(" No dimensioning scenario",id)
        scn = 'No dimensioning scenario for ' + id
        # print_cum_cube(id+" No dimensioning scenario ",IDAsorted)        
        if cube_result2file == True:
            print_cum_cube_file(id+" No dimensioning scenario ",IDAsorted,f_cube_result)        


    # if (InterpolationSuccess == True) and (WantPlot == True):
    if (WantPlot == True):
        ll = len(IDAsorted)
        ec = np.zeros([ll,2])
        i=0
        ec[i,1] = IDAsorted[-1][0] #the longest duration
        ec[i,0] = IDAsorted[-1][1] #Frequency for the longest duration
        for i in range(1,len(IDAsorted)):
            ec[i,1] = IDAsorted[ll-i-1][0]
            ec[i,0] = ec[i-1,0] + IDAsorted[ll-i-1][1]

        # plt.figure(figsize=(5.91, 3.15))
        CF = ec[1:,0]
        JFL = ec[1:,1]
        masscolor = 'tab:blue'
        fig,ax1 = plt.subplots()
        ax1.set_xlabel('Jet Impingement Duration [sec]')
        ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
        ax1.semilogy(JFL,CF,color=masscolor)
        ax1.set_ylim(top=5E-4,bottom=1E-6)
        ax1.set_xlim(left=0, right=3600)
        ax1.tick_params(axis='y',labelcolor=masscolor)
        ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
        ax1.annotate(scn,xy=(di,1.0E-4),xytext=(di,2E-4),horizontalalignment='left',verticalalignment='top',arrowprops = dict(facecolor='black',headwidth=4,width=2,headlength=4))
        # ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
        # ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
        # ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
        ax1.grid(True,which="major")

        if InterpolationSuccess:
            plt.title("Cube {:10s} - {:5s} fire from {:30} for {:8.1f} [sec]".format(id,weather_fire[6:],pv+"_"+hole+"_"+weather_fire[:5],di))            
        else:
            plt.title("Cube {:10s} - No dimensioning fire".format(id))            
        plt.tight_layout()
        plt.show()

        fig.savefig("{}.png".format(id))
        plt.close()

    if cube_result2file == True:
        f_cube_result.close()
# print_cum_cube(IDAsorted)

# for c in Cubes:
#     print("{:20s} {:8.1f} [min]".format(c,CubeImpingeDuration[c]/60))
# sys.stdout.close()